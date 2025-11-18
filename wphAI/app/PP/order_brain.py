# -*- coding: utf-8 -*-
r"""
Order Brain — Cheapest-only, Aggressive API Matching, Winner_Reason + Rabat columns
- Furnitori fitues = çmimi efektiv më i ulët (pas rabatit).
- Fallback: BARKOD/SIFRA → SIGNATURE → API_DOSE → SIGNATURE2 → API_NEAR.
- MATCH_STATUS: EXACT/FUZZY/NO_MATCH nga API+dozë(+formë) dhe paketimi.
- Porosi: IGNORE_ULAZ_IN_ORDER=True (ULAZ injorohet), TARGET_MODE="izlaz".
"""

import os, re, sys, traceback
from datetime import datetime, date
from math import ceil
from typing import Optional, Dict, Set, Tuple, List, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

# =================== CONFIG (URDHRI PA ULAZ) ===================
TARGET_MODE = "izlaz"          # "izlaz" | "mes_muj" | "max"
IGNORE_ULAZ_IN_ORDER = True    # True -> Available = Stanje; ULAZ injorohet
ROUND_TO_5_IF_GE_10 = True     # rrumbullakim në 5-she kur sasia >= 10

# =================== Logger & utils ===================
def log(msg: str) -> None:
    print(msg, flush=True)

def _norm(s: str) -> str:
    if s is None: return ""
    s = str(s)
    s = (s.replace("Š","S").replace("š","s").replace("Ž","Z").replace("ž","z")
           .replace("Ć","C").replace("ć","c").replace("Č","C").replace("č","c")
           .replace("Đ","DJ").replace("đ","dj"))
    return re.sub(r"[\s\W]+", "", s, flags=re.UNICODE).lower()

def pick_col_safe(df: pd.DataFrame, aliases: list[str]) -> Optional[str]:
    m = {_norm(c): c for c in df.columns}
    for a in aliases:
        key = _norm(a)
        if key in m: return m[key]
    return None

def s_clean(s: pd.Series) -> pd.Series:
    if s is None: return pd.Series([], dtype="string")
    return s.astype("string").fillna("").str.replace(r"\.0$", "", regex=True).str.strip()

def norm_name(s: str) -> str:
    x = str(s or "").upper()
    x = (x.replace("Š","S").replace("š","s").replace("Ž","Z").replace("ž","z")
           .replace("Ć","C").replace("ć","c").replace("Č","C").replace("č","c")
           .replace("Đ","DJ").replace("đ","dj"))
    x = re.sub(r"[^A-Z0-9\sxX.]", " ", x)
    x = re.sub(r"(?<=\d)[\s.](?=\d)", "", x)
    x = re.sub(r"\s+", " ", x).strip()
    return x

def norm_code(s: str) -> str:
    x = re.sub(r"[\s\-\_/\.]", "", str(s or ""))
    x = re.sub(r"[^A-Za-z0-9]", "", x).upper()
    x = x.lstrip("0")
    return x

# =================== ORDER POLICY HELPERS ===================
from app.core.policy import compute_order_qty  # central policy logic

# =================== Excel helpers ===================
def style_header(ws, row=1):
    hdr_fill = PatternFill("solid", fgColor="22313F")
    hdr_font = Font(color="FFFFFF", bold=True)
    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for c in ws[row]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(vertical="center", horizontal="center")
        c.border = border_thin

def auto_fit(ws, min_w=8, extra=2):
    for col in ws.columns:
        maxlen = 0
        idx = col[0].column if hasattr(col[0], "column") else col[0].col_idx
        for cell in col:
            s = "" if cell.value is None else str(cell.value)
            if len(s) > maxlen: maxlen = len(s)
        ws.column_dimensions[get_column_letter(idx)].width = max(min_w, maxlen + extra)

def add_table_safe(wb, ws, display_name: str, ref: str, style: str = "TableStyleMedium9") -> None:
    try:
        min_c, min_r, max_c, max_r = range_boundaries(ref)
    except Exception:
        log(f"[TABLE] Skip {display_name}: ref invalid {ref}"); return
    if max_r - min_r + 1 < 2 or max_c - min_c + 1 < 1:
        log(f"[TABLE] Skip {display_name}: range i vogël {ref}"); return
    headers = [ws.cell(row=min_r, column=c).value for c in range(min_c, max_c + 1)]
    if any(h is None or str(h).strip()=="" for h in headers):
        log(f"[TABLE] Skip {display_name}: header bosh"); return
    existing = set()
    for sh in wb.worksheets:
        for nm in getattr(sh, "tables", {}): existing.add(nm)
    name = re.sub(r"[^A-Za-z0-9_]", "_", display_name)
    if not re.match(r"^[A-Za-z_]", name): name = f"t_{name}"
    while name in existing: name += "_x"
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    try:
        ws.add_table(t)
    except Exception as e:
        log(f"[TABLE] Skip {name}: {e}")

# =================== Paths ===================
def parse_enddate_from_filename(fn: str) -> Optional[date]:
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", fn)
    if not m: return None
    d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100: y += 2000
    try: return date(y, mth, d)
    except ValueError: return None

def _desktop_root():
    try:
        return os.path.join(os.path.expanduser("~"), "Desktop", "OrderBrain_Exports")
    except Exception:
        return None

DEFAULT_OUT_ROOT = r"C:\Apoteka\OB\Outputs"

def get_outdir():
    candidates = []
    env_path = os.environ.get("OB_OUTDIR", "").strip()
    if env_path: candidates.append(env_path)
    desk = _desktop_root()
    if desk: candidates.append(desk)
    candidates.append(DEFAULT_OUT_ROOT)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates.append(os.path.join(script_dir, "Outputs"))
    candidates.append(os.path.join(os.getcwd(), "Outputs"))
    for root in candidates:
        try:
            day_dir = os.path.join(root, datetime.now().strftime("%Y-%m-%d"))
            os.makedirs(day_dir, exist_ok=True)
            return day_dir
        except Exception as e:
            log(f"[WARN] Outputs path not usable: {root} -> {e}")
    fallback = os.path.join(os.getcwd(), "Outputs", datetime.now().strftime("%Y-%m-%d"))
    os.makedirs(fallback, exist_ok=True)
    return fallback

def pastel_palette():
    return ["FFF4E5", "E6F7FF", "E9F7EF", "F6E8FF", "FFF0F0", "EAF2FF", "EFFFF9", "FFF9E6"]

# =================== API/DOSE parsing ===================
_FORM_MAP = {
    'tablet':'tablet','tablete':'tablet','tbl':'tablet','tab':'tablet','tabs':'tablet','comp':'tablet','tbl.':'tablet',
    'capsule':'capsule','caps':'capsule','cap':'capsule','kaps':'capsule','kapsule':'capsule','cps':'capsule',
    'inj':'injection','inj.':'injection','injeksion':'injection','injection':'injection','amp':'injection','ampoule':'injection',
    'sirup':'syrup','syrup':'syrup','shurup':'syrup','susp':'suspension','suspension':'suspension',
    'cream':'cream','krem':'cream','ung':'ointment','ointment':'ointment','pomade':'ointment',
    'spray':'spray','sprej':'spray','sol':'solution','solution':'solution','solucion':'solution','soln':'solution','sol.':'solution',
    'gel':'gel','drops':'drops','pika':'drops','drps':'drops'
}
_VENDOR_STOP = {'krka','teva','sandoz','actavis','bayer','pfizer','galenika','hemofarm','al','ratiopharm','generics','orion','stada','ferrer','abbvie','novartis','msd','gsk','lilly','roche','berlin','chemie'}
_UNIT_TOK = {'mg','mcg','g','kg','ml','l','iu','unit','%','mg/ml','mg/g','mcg/ml','mcg/g'}
_BRAND2API_BUILTIN = {
    "AMARYL":"GLIMEPIRIDE","BRUFEN":"IBUPROFEN","LAMICTAL":"LAMOTRIGINE","CRESTOR":"ROSUVASTATIN",
    "LORISTA":"LOSARTAN","PROSCAR":"FINASTERIDE","NOLVADEX":"TAMOXIFEN","VOLTAREN":"DICLOFENAC",
    "KETONAL":"KETOPROFEN","PANCEF":"CEFIXIME","ALCOHOL":"ETHANOL","HYDROGEN":"HYDROGEN PEROXIDE",
}

def normalize_name_string(name: str) -> str:
    s = str(name or "").lower()
    s = s.replace("µg","mcg").replace("μg","mcg")
    s = re.sub(r'[,\(\)\[\]]', ' ', s)
    s = re.sub(r'(?<=\d)x(?=\d)', ' x ', s)
    s = re.sub(r'(?<=[a-z])x(?=\d)', ' x ', s)
    s = re.sub(r'(?<=\d)x\b', ' x', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def _std_form(tok: str) -> Optional[str]:
    return _FORM_MAP.get(tok.strip().lower())

def _apply_brand_alias(text: str, brand_map: Dict[str,str]) -> str:
    t = text.upper()
    for brand, api in brand_map.items():
        t = re.sub(rf'\b{re.escape(brand)}\b', api, t)
    return t.title()

def make_main_key(api: Optional[str], dose_value: Optional[Union[float,int]], dose_unit: Optional[str], form: Optional[str]) -> tuple:
    return ((api or "").strip().lower(),
            dose_value if dose_value is not None else None,
            (dose_unit or "").strip().lower() or None,
            (form or "").strip().lower() or None)

def parse_api_dose_name(name: str, brand_map: Optional[Dict[str,str]]=None) -> dict:
    s = normalize_name_string(name)
    toks = s.split()
    form_found = None
    for i,t in enumerate(toks):
        f = _std_form(t)
        if f: form_found=f; toks[i]=''
    s_noform = " ".join([t for t in toks if t]).strip()

    packaging=None; dose_value=None; dose_unit=None

    mratio = re.search(r'(?P<d>\d+(?:[.,]\d+)?)\s*(?P<u>mcg|mg|g|iu)\s*/\s*(?P<den>\d+(?:[.,]\d+)?)\s*(?P<du>ml|g|l|kg)', s_noform)
    if mratio:
        d = float(mratio.group('d').replace(',','.')); u = mratio.group('u').lower()
        den = float(mratio.group('den').replace(',','.')); du = mratio.group('du').lower()
        if u=='g': d*=1000.0; u='mg'
        if u=='kg': d*=1_000_000.0; u='mg'
        if u=='iu': u='unit'
        if du=='l': den*=1000.0; du='ml'
        if du=='kg': den*=1000.0; du='g'
        dose_value = round(d/den, 6)
        dose_unit = f"{u}/{ 'ml' if du=='ml' else 'g'}"
        s_noform = (s_noform[:mratio.start()] + " " + s_noform[mratio.end():]).strip()

    if packaging is None:
        pm = re.search(r'(?P<pack>\d+)\s*x\b', s_noform)
        if pm:
            packaging = int(pm.group('pack'))
            s_noform = (s_noform[:pm.start()] + " " + s_noform[pm.end():]).strip()

    if dose_value is None:
        m1 = re.search(r'(?P<dose>\d+(?:[.,]\d+)?)\s*(?P<unit>mcg|mg|g|kg|ml|l|iu|unit)\b', s_noform)
        if m1:
            dose_value = float(m1.group('dose').replace(',','.'))
            dose_unit = m1.group('unit').lower()
            s_noform = (s_noform[:m1.start()] + " " + s_noform[m1.end():]).strip()

    if dose_value is None:
        mp = re.search(r'(?P<p>\d+(?:[.,]\d+)?)\s*%', s_noform)
        if mp:
            dose_value = float(mp.group('p').replace(',','.')); dose_unit = '%'
            s_noform = (s_noform[:mp.start()] + " " + s_noform[mp.end():]).strip()

    if dose_unit:
        if dose_unit=='g': dose_value*=1000.0; dose_unit='mg'
        elif dose_unit=='kg': dose_value*=1_000_000.0; dose_unit='mg'
        elif dose_unit=='l': dose_value*=1000.0; dose_unit='ml'
        elif dose_unit=='iu': dose_unit='unit'

    api_text = s_noform.strip("- ").strip()
    if brand_map: api_text = _apply_brand_alias(api_text, brand_map)
    api = api_text.strip()

    if dose_value is not None and float(dose_value).is_integer():
        dose_value = int(dose_value)

    return {'api': api, 'dose_value': dose_value, 'dose_unit': dose_unit,
            'form': form_found, 'packaging': packaging, 'original_name': name}

_STOP_FORM = set(_FORM_MAP.keys())
_STOP_MISC = {'kom','pcs','ks','x','plus','and','&','oral','sc','iv','im','sr','retard','forte','extra','pu','pu-q','pu-qs','pu-ks'}
_STOP = _STOP_FORM | _VENDOR_STOP | _UNIT_TOK | _STOP_MISC

def semantic_signature(name: str) -> str:
    s = normalize_name_string(name)
    s = re.sub(r'\b\d+([a-z%/]+)?\b', ' ', s)
    toks = re.split(r'[^a-z]+', s)
    toks = [t for t in toks if t and t not in _STOP]
    toks = sorted(set(toks))
    return " ".join(toks)

# =================== BRAND→API maps ===================
def _load_brand_alias_csv(path="drug_aliases.csv") -> Dict[str,str]:
    if not os.path.isfile(path): return {}
    try:
        df = pd.read_csv(path)
        df = df.rename(columns={c:c.lower() for c in df.columns})
        out = {}
        for _,r in df.iterrows():
            b = str(r.get('brand') or '').strip().upper()
            a = str(r.get('api') or '').strip().upper()
            if b and a: out[b]=a
        return out
    except Exception:
        return {}

def _save_alias_csv(d: Dict[str,str], path: str):
    try:
        if not d: return
        df = pd.DataFrame(sorted([(k,v) for k,v in d.items()]), columns=["brand","api"])
        df.to_csv(path, index=False, encoding="utf-8-sig")
        log(f"[ALIAS] Saved {len(d)} auto aliases -> {path}")
    except Exception as e:
        log(f"[ALIAS] Save failed: {e}")

def extract_brand_candidates(name: str, api_text: str) -> Optional[str]:
    s = normalize_name_string(name)
    tokens = [t for t in re.split(r'[^a-z]+', s) if t]
    api_tokens = set([t for t in re.split(r'[^a-z]+', api_text.lower()) if t])
    for tok in tokens:
        if tok in api_tokens: continue
        if tok in _STOP or tok in _VENDOR_STOP: continue
        if len(tok) < 3: continue
        return tok.upper()
    return None

def build_auto_brand_map(erp_names: pd.Series, furn_names: pd.Series, out_dir: str) -> Tuple[Dict[str,str], pd.DataFrame]:
    cand_counts: Dict[str, Dict[str,int]] = {}
    def feed(series: pd.Series):
        for name in series.dropna().astype(str):
            p = parse_api_dose_name(name, brand_map={})
            api = (p["api"] or "").strip()
            if not api: continue
            brand = extract_brand_candidates(name, api)
            if not brand: continue
            d = cand_counts.setdefault(brand, {})
            d[api.upper()] = d.get(api.upper(), 0) + 1
    feed(erp_names); feed(furn_names)
    auto: Dict[str,str] = {}
    rows: List[dict] = []
    for brand, apis in cand_counts.items():
        total = sum(apis.values())
        api_sorted = sorted(apis.items(), key=lambda kv: (-kv[1], kv[0]))
        winner_api, winner_cnt = (api_sorted[0] if api_sorted else ("", 0))
        if winner_cnt >= 3:
            auto[brand] = winner_api
        cand_str = " | ".join([f"{k}:{v}" for k,v in api_sorted[:6]])
        rows.append({"brand": brand, "winner_api": winner_api, "winner_count": winner_cnt,
                     "total": total, "candidates": cand_str})
    rep = pd.DataFrame(rows).sort_values(["winner_count","total"], ascending=[False, False])
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    _save_alias_csv(auto, os.path.join(out_dir, f"drug_aliases_auto_{ts}.csv"))
    _save_alias_csv(auto, "drug_aliases_auto.csv")
    return auto, rep

def build_brand_map(extra_csv: Optional[str]=None, auto_map: Optional[Dict[str,str]]=None) -> Dict[str,str]:
    m = dict(_BRAND2API_BUILTIN)
    if extra_csv and os.path.isfile(extra_csv): m.update(_load_brand_alias_csv(extra_csv))
    if os.path.isfile("drug_aliases.csv"): m.update(_load_brand_alias_csv("drug_aliases.csv"))
    if os.path.isfile("drug_aliases_auto.csv"):
        try:
            df = pd.read_csv("drug_aliases_auto.csv")
            for _,r in df.iterrows():
                b = str(r.get('brand') or '').strip().upper()
                a = str(r.get('api') or '').strip().upper()
                if b and a: m[b]=a
        except Exception:
            pass
    if auto_map: m.update(auto_map)
    return m

# =================== Supplier indexes (CHEAPEST ONLY + Rabat) ===================
def effective_price(df: pd.DataFrame, price_col: Optional[str], rabat_col: Optional[str]) -> Tuple[pd.Series, pd.Series]:
    price = pd.to_numeric(df[price_col], errors="coerce").fillna(0.0) if price_col else pd.Series([0.0]*len(df))
    rabat_pct = pd.Series([0.0]*len(df))
    if rabat_col and rabat_col in df.columns:
        rabat_pct = pd.to_numeric(df[rabat_col], errors="coerce").fillna(0.0)
        price = price * (1 - rabat_pct/100.0)
    return price, rabat_pct

def build_supplier_indexes(df_furn: pd.DataFrame, F_ART: str, F_SUP: str, F_SIFRA: Optional[str],
                           F_PRICE: Optional[str], F_RABAT: Optional[str],
                           brand_map: Dict[str,str]):
    eff_price, rabat_pct = effective_price(df_furn, F_PRICE, F_RABAT)
    catalog = pd.DataFrame({
        "Furnitor": df_furn[F_SUP].astype(str).str.upper().str.strip(),
        "SifraSup": df_furn[F_SIFRA].astype(str).fillna("").str.strip() if F_SIFRA else "",
        "Art": df_furn[F_ART].astype(str).fillna(""),
        "Price": eff_price,
        "RabatPct": rabat_pct.fillna(0.0)
    })
    parsed = catalog["Art"].map(lambda x: parse_api_dose_name(x, brand_map)).apply(pd.Series)
    for c in ["api","dose_value","dose_unit","form","packaging"]:
        catalog[c] = parsed[c]
    catalog["main_key"] = catalog.apply(lambda r: make_main_key(r["api"], r["dose_value"], r["dose_unit"], r["form"]), axis=1)
    catalog["Signature"]  = catalog["Art"].map(lambda x: re.sub(r'\s+',' ', x.upper().strip()))
    catalog["Signature2"] = catalog["Art"].map(semantic_signature)
    catalog["PriceValid"] = catalog["Price"].apply(lambda v: 0 if v and v>0 else 1)

    sig_best = catalog.sort_values(["Signature","PriceValid","Price"]).drop_duplicates(["Signature"], keep="first")
    SIG_GLOB = { r["Signature"]:(r["Furnitor"], r["SifraSup"], float(r["Price"]), float(r["RabatPct"])) for _,r in sig_best.iterrows() if r["Signature"] }

    sig2_best = catalog.sort_values(["Signature2","PriceValid","Price"]).drop_duplicates(["Signature2"], keep="first")
    SIG2_GLOB = { r["Signature2"]:(r["Furnitor"], r["SifraSup"], float(r["Price"]), float(r["RabatPct"])) for _,r in sig2_best.iterrows() if r["Signature2"] }

    API_GLOB: Dict[tuple, dict] = {}
    PKG_MAP: Dict[tuple, Set[int]] = {}
    API_PARTIAL: Dict[Tuple[str,str,str], List[Tuple[float,str,str,float,float]]] = {}
    for key, grp in catalog.groupby("main_key"):
        api, dose, unit, form = key
        if not api or dose is None or not unit: continue
        g2 = grp.sort_values(["PriceValid","Price"]).iloc[0]
        API_GLOB[key] = {"Furnitor":g2["Furnitor"], "Sifra":g2["SifraSup"], "Price":float(g2["Price"]), "RabatPct": float(g2["RabatPct"])}
        PKG_MAP[key] = set(int(x) for x in grp["packaging"].dropna().astype(int).tolist())
        base = (api, unit, form or "")
        lst = API_PARTIAL.setdefault(base, [])
        for _, rr in grp.iterrows():
            if pd.isna(rr["dose_value"]): continue
            lst.append((float(rr["dose_value"]), rr["Furnitor"], rr["SifraSup"], float(rr["Price"]), float(rr["RabatPct"])))
        API_PARTIAL[base] = sorted(lst, key=lambda t: (0 if t[3] and t[3]>0 else 1, t[3]))

    log(f"[INDEX] Built (cheapest-only): SIG={len(SIG_GLOB)} SIG2={len(SIG2_GLOB)} API={len(API_GLOB)}")
    return SIG_GLOB, SIG2_GLOB, API_GLOB, PKG_MAP, API_PARTIAL, catalog

# =================== Fallbacks ===================
def _near_tolerance(unit: Optional[str], form: Optional[str]) -> float:
    u = (unit or "").lower(); f = (form or "").lower()
    if u.endswith("/ml") or u.endswith("/g") or f in {"solution","syrup","suspension","drops"}: return 0.15
    if f in {"tablet","capsule"}: return 0.05
    return 0.10

def apply_enhanced_fallbacks(df_review: pd.DataFrame, supplier_indexes, brand_map: Dict[str,str]) -> pd.DataFrame:
    SIG, SIG2, API, PKG, PARTIAL, _catalog = supplier_indexes
    before = int((df_review["Match_method"] == "UNMATCHED").sum())
    erp_sig  = df_review["Artikal"].astype(str).str.upper().str.strip()
    erp_sig2 = df_review["Artikal"].astype(str).map(semantic_signature)
    erp_parsed = df_review["Artikal"].astype(str).map(lambda x: parse_api_dose_name(x, brand_map)).apply(pd.Series)
    erp_main_key = erp_parsed.apply(lambda r: make_main_key(r["api"], r["dose_value"], r["dose_unit"], r["form"]), axis=1)

    improved = 0; near_hits = 0
    mask = df_review["Match_method"].astype(str).eq("UNMATCHED")
    for i in df_review.index[mask]:
        hit = SIG.get(erp_sig.iat[i])
        if hit:
            furn, _, price, rab = hit
            df_review.at[i,"Furnitor_best"]=furn; df_review.at[i,"Cmim_best"]=price
            df_review.at[i,"Match_method"]="SIGNATURE"; df_review.at[i,"Winner_Reason"]="PRICE_MIN_SIGNATURE"
            df_review.at[i,"__Rabat_tmp__"]=rab; improved += 1; continue

        info = API.get(erp_main_key.iat[i])
        if info:
            df_review.at[i,"Furnitor_best"]=info["Furnitor"]; df_review.at[i,"Cmim_best"]=info["Price"]
            df_review.at[i,"Match_method"]="API_DOSE"; df_review.at[i,"Winner_Reason"]="PRICE_MIN_API_DOSE"
            df_review.at[i,"__Rabat_tmp__"]=info.get("RabatPct",0.0); improved += 1; continue

        hit2 = SIG2.get(erp_sig2.iat[i])
        if hit2:
            furn, _, price, rab = hit2
            df_review.at[i,"Furnitor_best"]=furn; df_review.at[i,"Cmim_best"]=price
            df_review.at[i,"Match_method"]="SIGNATURE2"; df_review.at[i,"Winner_Reason"]="PRICE_MIN_SIGNATURE2"
            df_review.at[i,"__Rabat_tmp__"]=rab; improved += 1; continue

        api, dose, unit, form = erp_main_key.iat[i]
        if api and dose is not None and unit:
            base = (api, unit, form or "")
            lst = PARTIAL.get(base)
            if lst:
                tol = _near_tolerance(unit, form)
                best = None; best_rel = 1e9
                for dval, furn, sifra, price, rab in lst:
                    if float(dose) == 0: continue
                    rel = abs(dval - float(dose)) / max(abs(float(dose)), 1e-9)
                    if rel <= tol and rel < best_rel:
                        best = (furn, price, rab, rel); best_rel = rel
                if best:
                    furn, price, rab, rel = best
                    df_review.at[i,"Furnitor_best"]=furn; df_review.at[i,"Cmim_best"]=price
                    df_review.at[i,"Match_method"]="API_NEAR"
                    df_review.at[i,"Winner_Reason"]=f"PRICE_MIN_API_NEAR(<= {int(tol*100)}%)"
                    df_review.at[i,"__Rabat_tmp__"]=rab; improved += 1; near_hits += 1
                    continue

    after = int((df_review["Match_method"] == "UNMATCHED").sum())
    log(f"[FALLBACK+] UNMATCHED para: {before} -> pas: {after}  (zbritje: {before-after}, API_NEAR={near_hits})")
    return df_review

def compute_match_status_api_dose(df_review: pd.DataFrame, PKG_MAP: Dict[tuple, Set[int]], brand_map: Dict[str,str]) -> pd.Series:
    parsed = df_review["Artikal"].astype(str).map(lambda x: parse_api_dose_name(x, brand_map)).apply(pd.Series)
    keys = parsed.apply(lambda r: make_main_key(r["api"], r["dose_value"], r["dose_unit"], r["form"]), axis=1)
    packs = parsed["packaging"].fillna(1).astype(int)
    st = []; ex=fz=no=0
    for k,p in zip(keys, packs):
        if not k or (k[0] in (None,"") or k[1] is None or k[2] in (None,"")):
            st.append("NO_MATCH"); no+=1; continue
        pkgs = PKG_MAP.get(k)
        if pkgs is None:
            st.append("NO_MATCH"); no+=1
        else:
            if p in pkgs: st.append("EXACT"); ex+=1
            else: st.append("FUZZY"); fz+=1
    log(f"[API-MATCH] Status -> EXACT:{ex} | FUZZY:{fz} | NO_MATCH:{no}")
    return pd.Series(st, index=df_review.index, dtype="string")

# =================== Rabat columns ===================
def add_rabat_columns(df_review: pd.DataFrame,
                      furn_tbl: pd.DataFrame,
                      catalog: pd.DataFrame,
                      brand_map: Dict[str,str]) -> pd.DataFrame:
    rb_bark_sup = {(b, s): r for b,s,r in zip(furn_tbl["Barkod"], furn_tbl["Furnitor"], furn_tbl["RabatPct"])}
    rb_sif_sup  = {(c, s): r for c,s,r in zip(furn_tbl["Sifra"],  furn_tbl["Furnitor"], furn_tbl["RabatPct"])}

    cat_by_sig   = {k: g for k, g in catalog.groupby("Signature", dropna=False)}
    cat_by_sig2  = {k: g for k, g in catalog.groupby("Signature2", dropna=False)}
    cat_by_key   = {k: g for k, g in catalog.groupby("main_key", dropna=False)}

    def near_group(parsed_row):
        api, dose, unit, form = parsed_row["api"], parsed_row["dose_value"], parsed_row["dose_unit"], parsed_row["form"]
        if not api or dose is None or not unit: return None
        base_g = catalog[(catalog["api"].str.lower()==str(api).lower()) &
                         (catalog["dose_unit"].astype(str).str.lower()==str(unit).lower()) &
                         (catalog["form"].astype(str).str.lower()==str(form or "").lower())]
        tol = _near_tolerance(unit, form)
        return base_g[ (base_g["dose_value"].astype(float)-float(dose)).abs()/max(abs(float(dose)),1e-9) <= tol ]

    rb_best = []; rb_oth_max = []; rb_oth_avg = []
    parsed_all = df_review["Artikal"].astype(str).map(lambda x: parse_api_dose_name(x, brand_map)).apply(pd.Series)

    for i, row in df_review.iterrows():
        sup = str(row.get("Furnitor_best") or "").upper().strip()
        mth = str(row.get("Match_method") or "")
        rbest = None; rother_max = 0.0; rother_avg = 0.0

        if mth in ("BARKOD","SIFRA"):
            if mth=="BARKOD":
                code = str(row.get("Barkod_actual") or "").strip()
                grp = furn_tbl[furn_tbl["Barkod"]==code] if code else pd.DataFrame()
                if code: rbest = rb_bark_sup.get((code, sup), None)
            else:
                code = str(row.get("Sifra_actual") or "").strip()
                grp = furn_tbl[furn_tbl["Sifra"]==code] if code else pd.DataFrame()
                if code: rbest = rb_sif_sup.get((code, sup), None)
            if code and not grp.empty:
                others = grp.loc[grp["Furnitor"].ne(sup), "RabatPct"]
                rother_max = float(others.max()) if not others.empty else 0.0
                rother_avg = float(others.mean()) if not others.empty else 0.0

        elif mth in ("SIGNATURE","SIGNATURE2",""):
            sig = row["Artikal"].upper().strip() if mth=="SIGNATURE" else semantic_signature(row["Artikal"])
            g = cat_by_sig.get(sig) if mth=="SIGNATURE" else cat_by_sig2.get(sig)
            if g is not None and not g.empty:
                g_sorted = g.sort_values(["PriceValid","Price"])
                rbest = float(g_sorted.iloc[0]["RabatPct"])
                others = g_sorted.loc[g_sorted["Furnitor"]!=sup,"RabatPct"] if sup else g_sorted["RabatPct"].iloc[1:]
                rother_max = float(others.max()) if not others.empty else 0.0
                rother_avg = float(others.mean()) if not others.empty else 0.0

        elif mth in ("API_DOSE","API_NEAR"):
            pr = parsed_all.loc[i]
            if mth=="API_DOSE":
                key = make_main_key(pr["api"], pr["dose_value"], pr["dose_unit"], pr["form"])
                g = cat_by_key.get(key)
            else:
                g = near_group(pr)
            if g is not None and not g.empty:
                g_sorted = g.sort_values(["PriceValid","Price"])
                rbest = float(g_sorted.iloc[0]["RabatPct"])
                others = g_sorted.loc[g_sorted["Furnitor"]!=sup,"RabatPct"] if sup else g_sorted["RabatPct"].iloc[1:]
                rother_max = float(others.max()) if not others.empty else 0.0
                rother_avg = float(others.mean()) if not others.empty else 0.0

        if rbest is None:
            rbest = float(row.get("__Rabat_tmp__", 0.0))

        rb_best.append(float(rbest or 0.0))
        rb_oth_max.append(float(rother_max or 0.0))
        rb_oth_avg.append(float(rother_avg or 0.0))

    df_review = df_review.copy()
    df_review["Rabat_best"] = rb_best
    df_review["Rabat_others_max"] = rb_oth_max
    df_review["Rabat_others_avg"] = rb_oth_avg
    if "__Rabat_tmp__" in df_review.columns:
        df_review.drop(columns=["__Rabat_tmp__"], inplace=True, errors="ignore")
    return df_review

def sup_id_for_row(row, sup_u, sup_bark_map, sup_sig_map):
    # Pse komentohet: fallback i fundit është barkodi, për t’ju krijuar porosi edhe kur s’ka SifraSup.
    bark = str(row.get("Barkod_actual") or "").strip()
    if bark and sup_u in sup_bark_map:
        hit = sup_bark_map[sup_u].get(bark)
        if hit: return hit
    sig = norm_name(str(row.get("Artikal") or ""))
    if sup_u in sup_sig_map:
        hit = sup_sig_map[sup_u].get(sig)
        if hit: return hit
    return bark

# =================== Main ===================
def main():
    log("[INFO] Running Order Brain (Cheapest-only, Winner_Reason, Rabat avg/max)…")
    # ---- CSV CONFIG ----
    EXPORT_ALL_SUPPLIERS = False
    SUPPLIERS = ["PHOENIX", "VEGA", "SOPHARMA"]
    ONLY_DIGIT_NAMES = True
    STRICT_SIFRA_ONLY = False
    CSV_DELIM = ";"
    CSV_HEADER = True
    DECIMAL_COMMA = False

    out_dir = get_outdir()
    log(f"[OK] Outputs directory: {out_dir}")

    # Files
    promet_curr = "PROMET ARTIKLA.xlsx"
    cands = [f for f in os.listdir(".") if f.lower().startswith("promet artikla ") and f.lower().endswith(".xlsx")]
    furn_file = "FURNITORET.xlsx"
    if not os.path.isfile(promet_curr): log("[ERR] Mungon 'PROMET ARTIKLA.xlsx'"); sys.exit(2)
    if not cands: log("[ERR] Mungon 'PROMET ARTIKLA <date>.xlsx'"); sys.exit(3)
    cands.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    promet_dated = cands[0]
    if not os.path.isfile(furn_file): log("[ERR] Mungon FURNITORET.xlsx"); sys.exit(4)
    log(f"[INFO] Using: {promet_curr} | {promet_dated} | {furn_file}")

    # Load
    try:
        df_curr = pd.read_excel(promet_curr, engine="openpyxl"); log(f"[OK] PROMET current rows: {len(df_curr)}")
    except Exception:
        log("[ERR] Leximi i PROMET (pa datë) dështoi:"); log(traceback.format_exc()); sys.exit(5)
    try:
        df_dated = pd.read_excel(promet_dated, engine="openpyxl"); log(f"[OK] PROMET dated rows: {len(df_dated)}")
    except Exception:
        log("[ERR] Leximi i PROMET (me datë) dështoi:"); log(traceback.format_exc()); sys.exit(6)
    try:
        df_furn = pd.read_excel(furn_file, engine="openpyxl"); log(f"[OK] FURN rows: {len(df_furn)}")
    except Exception:
        log("[ERR] Leximi i FURN dështoi:"); log(traceback.format_exc()); sys.exit(7)

    # Column maps
    P_SIFRA = pick_col_safe(df_curr, ["Sifra","Šifra","Sifra Artikla","Sifra_artikla","Sifra_art","Code","SKU"])
    P_BARK  = pick_col_safe(df_curr, ["Barkod","Barcode"])
    P_ART   = pick_col_safe(df_curr, ["Artikal","Artikull","Emertim","Naziv","Opis","Name","Product","Item","Lek","Proizvod"]) or df_curr.columns[0]
    P_STAN  = pick_col_safe(df_curr, ["Stanje","Stock","Stoqe"]) or df_curr.columns[1]
    P_MIN   = pick_col_safe(df_curr, ["Min_zal","Min. zal.","MinZal","Min__zal"])
    P_IZL   = pick_col_safe(df_curr, ["Izlaz","Sales"])
    P_ULZ   = pick_col_safe(df_curr, ["Ulaz","Inbound","Incoming"])
    P_MOQ   = pick_col_safe(df_curr, ["MOQ","Paket","Pack","Pak"])
    log(f"[MAP] PROMET current -> Sifra={P_SIFRA}, Barkod={P_BARK}, Artikal={P_ART}, Stanje={P_STAN}, Min_zal={P_MIN}, Izlaz={P_IZL}, Ulaz={P_ULZ}, MOQ={P_MOQ}")

    D_SIFRA = pick_col_safe(df_dated, ["Sifra","Šifra","Sifra Artikla","Sifra_artikla","Sifra_art","Code","SKU"])
    D_BARK  = pick_col_safe(df_dated, ["Barkod","Barcode"])
    D_ART   = pick_col_safe(df_dated, ["Artikal","Artikull","Emertim","Naziv","Opis","Name","Product","Item","Lek","Proizvod"]) or df_dated.columns[0]
    D_IZL   = pick_col_safe(df_dated, ["Izlaz","Sales"])
    if not D_IZL: log("[ERR] PROMET (me datë) pa Izlaz/Sales"); sys.exit(8)

    F_SUP   = pick_col_safe(df_furn, ["Furnitor","Dobavljac","Dobavljač","Supplier","Furnitor_best"])
    F_BARK  = pick_col_safe(df_furn, ["Barkod","Barcode","barkod"])
    F_SIFRA = pick_col_safe(df_furn, ["Sifra","Šifra","Sifra Artikla","sifra artikla","sifra_artikla","sifra_art","Code","SKU"])
    F_ART   = pick_col_safe(df_furn, ["Artikal","Artikull","Emertim","Naziv","Opis","Name","Product","Item","Lek","Proizvod"]) or df_furn.columns[0]
    F_PRICE = pick_col_safe(df_furn, ["Cmim","Price","Cena","vp_cena","vpcena","vp cena","VP CENA","VP cena"]) or df_furn.columns[-1]
    F_RABAT = pick_col_safe(df_furn, ["Rabat","Discount","rabat"])
    log(f"[MAP] FURN -> Furnitor={F_SUP}, Barkod={F_BARK}, Sifra={F_SIFRA}, Artikal={F_ART}, Price={F_PRICE}, Rabat={F_RABAT}")

    if not F_SUP:
        df_furn["__SUP__"] = "UNKNOWN"; F_SUP="__SUP__"; log("[WARN] Furnitor mungon; përdor 'UNKNOWN'.")

    # BRAND alias
    auto_alias, alias_report = build_auto_brand_map(df_curr[P_ART], df_furn[F_ART], out_dir=out_dir)
    brand_map = build_brand_map(auto_map=auto_alias)
    log(f"[ALIAS] Built brand map: total={len(brand_map)} (auto={len(auto_alias)})")

    # FURN normalized (cheapest-only + rabat)
    price = pd.to_numeric(df_furn[F_PRICE], errors="coerce").fillna(0.0)
    rabat_pct = pd.to_numeric(df_furn[F_RABAT], errors="coerce").fillna(0.0) if F_RABAT else pd.Series([0.0]*len(df_furn))
    if F_RABAT:
        price = price * (1 - rabat_pct/100.0)

    furn_tbl = pd.DataFrame({
        "Furnitor": df_furn[F_SUP].astype(str).str.upper().str.strip(),
        "Barkod":   s_clean(df_furn[F_BARK]) if F_BARK else pd.Series([""]*len(df_furn)),
        "Sifra":    s_clean(df_furn[F_SIFRA]) if F_SIFRA else pd.Series([""]*len(df_furn)),
        "Price":    price,
        "RabatPct": rabat_pct.fillna(0.0)
    })
    furn_tbl["Sifra"] = furn_tbl["Sifra"].map(norm_code)
    furn_tbl["Signature"] = df_furn[F_ART].astype(str).map(norm_name)
    furn_tbl["PriceValid"] = furn_tbl["Price"].apply(lambda v: 0 if v and v>0 else 1)

    # Global cheapest maps (BARKOD/SIFRA → (Furnitor, Price, Rabat))
    MAP_BARK, MAP_SIF = {}, {}
    if F_BARK:
        tmp = (furn_tbl[furn_tbl["Barkod"]!=""].sort_values(["Barkod","PriceValid","Price"]).groupby("Barkod", as_index=False).first())
        MAP_BARK = dict(zip(tmp["Barkod"], zip(tmp["Furnitor"], tmp["Price"], tmp["RabatPct"])))
    if F_SIFRA:
        tmp = (furn_tbl[furn_tbl["Sifra"]!=""].sort_values(["Sifra","PriceValid","Price"]).groupby("Sifra", as_index=False).first())
        MAP_SIF = dict(zip(tmp["Sifra"], zip(tmp["Furnitor"], tmp["Price"], tmp["RabatPct"])))
    log(f"[FURN] Index sizes -> Barkod:{len(MAP_BARK)} Sifra:{len(MAP_SIF)}")

    # Mes Muj (YTD→mujor)
    end_dt = parse_enddate_from_filename(promet_dated) or date.today()
    start_dt = date(end_dt.year,1,1)
    days = max((end_dt - start_dt).days + 1, 1)
    factor = 30.0 / days
    log(f"[MES] period {start_dt}..{end_dt} days={days} factor={factor:.6f}")

    MES_BARK, MES_SIF = {}, {}
    if D_BARK:
        g = df_dated.dropna(subset=[D_IZL]).copy()
        g[D_IZL] = pd.to_numeric(g[D_IZL], errors="coerce").fillna(0.0)
        t = g.groupby(D_BARK, dropna=True)[D_IZL].sum().reset_index()
        MES_BARK = dict(zip(s_clean(t[D_BARK]), (t[D_IZL]*factor).round(6)))
    if D_SIFRA:
        g = df_dated.dropna(subset=[D_IZL]).copy()
        g[D_IZL] = pd.to_numeric(g[D_IZL], errors="coerce").fillna(0.0)
        t = g.groupby(D_SIFRA, dropna=True)[D_IZL].sum().reset_index()
        MES_SIF = dict(zip(s_clean(t[D_SIFRA]).map(norm_code), (t[D_IZL]*factor).round(6)))

    # PROMET cleaned
    prom_bark = s_clean(df_curr[P_BARK]) if P_BARK else pd.Series([""]*len(df_curr), dtype="string")
    prom_sifra_raw = s_clean(df_curr[P_SIFRA]) if P_SIFRA else pd.Series([""]*len(df_curr), dtype="string")
    prom_sifra = prom_sifra_raw.map(norm_code)

    # ---------------- Review rows ----------------
    rows = []
    n = len(df_curr)
    for idx, r in df_curr.iterrows():
        if idx % 250 == 0: log(f"[INFO] Processing {idx}/{n} ...")
        sifra = prom_sifra.iat[idx] if P_SIFRA else ""
        bark  = prom_bark.iat[idx] if P_BARK else ""
        name  = str(r[P_ART]).strip() if pd.notna(r.get(P_ART)) else ""
        stan  = float(r[P_STAN]) if pd.notna(r.get(P_STAN)) else 0.0
        minz  = float(r[P_MIN]) if P_MIN and pd.notna(r.get(P_MIN)) else 0.0
        izl   = float(r[P_IZL]) if P_IZL and pd.notna(r.get(P_IZL)) else 0.0
        ulz   = float(r[P_ULZ]) if P_ULZ and pd.notna(r.get(P_ULZ)) else 0.0
        moq   = int(r[P_MOQ]) if P_MOQ and pd.notna(r.get(P_MOQ)) else 0

        if bark and bark in MES_BARK: mes_muj = float(MES_BARK[bark])
        elif sifra and sifra in MES_SIF: mes_muj = float(MES_SIF[sifra])
        else: mes_muj = 0.0

        furnitor_best, price_best, method, rab_best_tmp, reason = "", 0.0, "UNMATCHED", 0.0, ""
        if bark and bark in MAP_BARK:
            furnitor_best, price_best, rab_best_tmp = MAP_BARK[bark]; method="BARKOD"; reason="PRICE_MIN_BARKOD"
        elif sifra and sifra in MAP_SIF:
            furnitor_best, price_best, rab_best_tmp = MAP_SIF[sifra]; method="SIFRA"; reason="PRICE_MIN_SIFRA"

        # ---- ORDER POLICY (ULAZ ignored, target = IZLAZ by default) ----
        target, available_calc, por_calc, por_final = compute_order_qty(
            stanje=stan, ulaz=ulz, mes_muj=mes_muj, izlaz=izl,
            min_zal=minz, moq=moq,
            target_mode=TARGET_MODE,
            ignore_ulaz=IGNORE_ULAZ_IN_ORDER,
            round_to_5_if_ge_10=ROUND_TO_5_IF_GE_10
        )

        needs = "Po" if por_final > 0 else "Jo"
        value = round(por_final * float(price_best or 0.0), 2)

        rows.append({
            "Mes_Muj": round(float(mes_muj or 0), 2),
            "Stanje": float(stan),
            "Artikal": name,
            "Porosit_final": int(por_final),
            "Izlaz": float(izl),
            "Furnitor_best": furnitor_best,
            "Needs_Order": needs,
            "value": value,
            "Barkod": method,
            "Cmim_best": float(price_best or 0.0),
            "Target_Stock": round(float(target or 0), 2),
            "Porosit_calc": round(float(por_calc or 0), 2),
            "Available_calc": round(float(available_calc or 0), 2),
            "Ulaz": float(ulz),
            "Barkod_actual": bark if bark else "",
            "Sifra_actual": sifra if sifra else "",
            "Match_method": method,
            "Winner_Reason": reason,
            "Factor_used": round(float(factor), 6),
            "MM_div_Izlaz": round((float(mes_muj)/float(izl)) if float(izl)>0 else 0, 6),
            "__Rabat_tmp__": float(rab_best_tmp or 0.0),
        })
    df_review = pd.DataFrame(rows)
    log(f"[OK] Review built: {len(df_review)} rows")

    # Supplier indexes (cheapest-only) + catalog
    sig, sig2, api_glob, pkg_map, api_partial, catalog = build_supplier_indexes(
        df_furn, F_ART, F_SUP, F_SIFRA, F_PRICE, F_RABAT, brand_map
    )

    # Fallbacks
    df_review = apply_enhanced_fallbacks(df_review, (sig, sig2, api_glob, pkg_map, api_partial, catalog), brand_map)

    # MATCH_STATUS (API/DOSE)
    df_review["MATCH_STATUS"] = compute_match_status_api_dose(df_review, pkg_map, brand_map)

    # Rabat columns
    df_review = add_rabat_columns(df_review, furn_tbl, catalog, brand_map)

    # ---------------- KPI ----------------
    total_items       = len(df_review)
    items_to_order    = int((df_review["Porosit_final"] > 0).sum())
    neg_stock_count   = int((df_review["Stanje"] < 0).sum())
    total_order_value = float(df_review.loc[df_review["Porosit_final"] > 0, "value"].sum())
    high_price_count  = int((df_review["Cmim_best"] > 1000).sum())
    unmatched_count   = int((df_review["Match_method"] == "UNMATCHED").sum())
    api_exact_count   = int((df_review["MATCH_STATUS"] == "EXACT").sum())
    api_fuzzy_count   = int((df_review["MATCH_STATUS"] == "FUZZY").sum())
    api_no_count      = int((df_review["MATCH_STATUS"] == "NO_MATCH").sum())

    # ---------------- Workbook ----------------
    wb = Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])

    # Home
    ws_home = wb.create_sheet("Home", 0)
    ws_home["A1"] = "ORDER BRAIN — NAVIGATION"; ws_home["A1"].font = Font(size=16, bold=True)
    ws_home["A3"] = "Kliko për seksionet & pamjet e shpejta:"; ws_home["A3"].font = Font(bold=True)
    for i, name in enumerate(
        ["Dashboard", "Review", "Summary", "MatchStats", "API_MatchStats", "Brand_Alias_Report",
         "Watchlist", "Top20", "Exceptions", "ABC_Items", "ABC_Suppliers", "Legend",
         "View_OrderOnly", "View_Unmatched", "View_HighValue"], start=5):
        ws_home.cell(row=i, column=1, value=f'=HYPERLINK("#\'{name}\'!A1","→ {name}")')
    auto_fit(ws_home)

    # Dashboard
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = "Order Brain — Dashboard"; ws_dash["A1"].font = Font(size=16, bold=True)
    ws_dash["A2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    kpis = [("Total items", total_items), ("Items to order", items_to_order), ("Negative stock", neg_stock_count),
            ("Unmatched", unmatched_count), ("API EXACT", api_exact_count), ("API FUZZY", api_fuzzy_count),
            ("API NO_MATCH", api_no_count), ("Total order value", total_order_value), ("Price > 1000", high_price_count)]
    r0 = 4
    border_thin = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                         top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    for i,(lab,val) in enumerate(kpis, start=r0):
        ws_dash.cell(row=i, column=1, value=lab).font = Font(bold=True)
        ws_dash.cell(row=i, column=2, value=val).font = Font(bold=True)
        col = "E7F4EA" if lab in ("Total items","Items to order","Total order value","API EXACT") else ("FFF5E5" if lab in ("Negative stock","Unmatched","API NO_MATCH") else "EAF2FF")
        for c in (1,2):
            ws_dash.cell(row=i, column=c).fill = PatternFill("solid", fgColor=col)
            ws_dash.cell(row=i, column=c).border = border_thin
    ws_dash.cell(row=r0+6, column=2).number_format = "#,##0.00"
    ws_dash.cell(row=12, column=1, value="Orders by Supplier").font = Font(bold=True)

    orders = (df_review[df_review["Porosit_final"] > 0]
              .groupby("Furnitor_best")
              .agg(Artikuj=("Artikal","count"), Sasi=("Porosit_final","sum"), Vlere=("value","sum"))
              .reset_index().sort_values("Vlere", ascending=False))
    ws_dash.append(["Supplier","Items","Qty","Value"]); style_header(ws_dash, row=13)
    for _, row in orders.iterrows():
        ws_dash.append([row["Furnitor_best"], int(row["Artikuj"]), int(row["Sasi"]), float(row["Vlere"])])
    if len(orders) > 0:
        dash_last = 13 + len(orders)
        add_table_safe(wb, ws_dash, "tblOrders", f"A13:D{dash_last}")
        ws_dash.conditional_formatting.add(f"D14:D{dash_last}", DataBarRule(start_type="min", end_type="max", color="9CA3AF"))
        try:
            chart = BarChart(); chart.title = "Order Value by Supplier"
            data = Reference(ws_dash, min_col=4, min_row=13, max_row=dash_last)
            cats = Reference(ws_dash, min_col=1, min_row=14, max_row=dash_last)
            chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
            chart.height = 9; chart.width = 18; ws_dash.add_chart(chart, "F13")
        except Exception: pass
    auto_fit(ws_dash)

    # Review
    ws_rev = wb.create_sheet("Review")
    headers = ["Mes_Muj","Stanje","Artikal","Porosit_final","Izlaz","Furnitor_best","Needs_Order",
               "value","Barkod","Cmim_best","Target_Stock","Porosit_calc","Available_calc","Ulaz",
               "Barkod_actual","Sifra_actual","Match_method","MATCH_STATUS","Winner_Reason",
               "Rabat_best","Rabat_others_max","Rabat_others_avg",
               "Factor_used","MM_div_Izlaz"]
    ws_rev.append(headers); style_header(ws_rev, row=1)
    for _, row in df_review.iterrows(): ws_rev.append([row.get(c,"") for c in headers])
    last_rev = ws_rev.max_row
    for r in range(2,last_rev+1):
        ws_rev.cell(row=r, column=headers.index("value")+1).number_format = "#,##0.00"
        ws_rev.cell(row=r, column=headers.index("Cmim_best")+1).number_format = "#,##0.00"
        ws_rev.cell(row=r, column=headers.index("Target_Stock")+1).number_format = "#,##0.##"
        ws_rev.cell(row=r, column=headers.index("Porosit_calc")+1).number_format = "#,##0.##"
        ws_rev.cell(row=r, column=headers.index("Rabat_best")+1).number_format = "0.00%"
        ws_rev.cell(row=r, column=headers.index("Rabat_others_max")+1).number_format = "0.00%"
        ws_rev.cell(row=r, column=headers.index("Rabat_others_avg")+1).number_format = "0.00%"
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    c_st = headers.index("Stanje")+1; c_need = headers.index("Needs_Order")+1; c_val = headers.index("value")+1
    ws_rev.conditional_formatting.add(f"{get_column_letter(c_st)}2:{get_column_letter(c_st)}{last_rev}", CellIsRule(operator="lessThan", formula=["0"], fill=red))
    ws_rev.conditional_formatting.add(f"{get_column_letter(c_need)}2:{get_column_letter(c_need)}{last_rev}", FormulaRule(formula=[f'INDIRECT("{get_column_letter(c_need)}"&ROW())="Po"'], fill=green))
    ws_rev.conditional_formatting.add(f"{get_column_letter(c_val)}2:{get_column_letter(c_val)}{last_rev}", DataBarRule(start_type="min", end_type="max", color="9CA3AF"))
    add_table_safe(wb, ws_rev, "tblReview", f"A1:{get_column_letter(len(headers))}{last_rev}")
    ws_rev.freeze_panes = "A2"; ws_rev.column_dimensions[get_column_letter(headers.index("Artikal")+1)].width = 46
    for r in range(2,last_rev+1):
        ws_rev.cell(row=r, column=headers.index("Artikal")+1).alignment = Alignment(wrap_text=True, vertical="center")
    auto_fit(ws_rev)

    # Summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Dobavljac","Value","Share","CumShare","Class"]); style_header(ws_sum, row=1)
    total_val = df_review["value"].sum() or 1.0
    sups = df_review.groupby("Furnitor_best")["value"].sum().reset_index().sort_values("value", ascending=False)
    sups["Share"] = sups["value"]/total_val; sups["CumShare"]=sups["Share"].cumsum()
    sups["Class"] = sups["CumShare"].map(lambda p: "A" if p<=0.80 else ("B" if p<=0.95 else "C"))
    for _, r in sups.iterrows():
        ws_sum.append([r["Furnitor_best"], float(r["value"]), float(r["Share"]), float(r["CumShare"]), r["Class"]])
    for r in range(2, ws_sum.max_row + 1):
        ws_sum.cell(row=r, column=2).number_format = "#,##0.00"
        ws_sum.cell(row=r, column=3).number_format = "0.00%"
        ws_sum.cell(row=r, column=4).number_format = "0.00%"
    add_table_safe(wb, ws_sum, "tblSummary", f"A1:E{ws_sum.max_row}"); auto_fit(ws_sum)

    # MatchStats
    ws_ms = wb.create_sheet("MatchStats")
    stats = (df_review.groupby("Match_method").agg(Rreshta=("Artikal","count"), Qty=("Porosit_final","sum"), Vlere=("value","sum"))
             .reset_index().sort_values("Rreshta", ascending=False))
    ws_ms.append(["Match_method","Rreshta","Qty","Vlere"]); style_header(ws_ms, row=1)
    for _, row in stats.iterrows(): ws_ms.append([row["Match_method"], int(row["Rreshta"]), int(row["Qty"]), float(row["Vlere"])])
    for r in range(2, ws_ms.max_row + 1): ws_ms.cell(row=r, column=4).number_format = "#,##0.00"
    add_table_safe(wb, ws_ms, "tblMatchStats", f"A1:D{ws_ms.max_row}"); auto_fit(ws_ms)

    # API_MatchStats
    ws_ams = wb.create_sheet("API_MatchStats")
    stats2 = (df_review.groupby("MATCH_STATUS").agg(Rreshta=("Artikal","count"), Qty=("Porosit_final","sum"), Vlere=("value","sum"))
              .reset_index().sort_values("Rreshta", ascending=False))
    ws_ams.append(["MATCH_STATUS","Rreshta","Qty","Vlere"]); style_header(ws_ams, row=1)
    for _, row in stats2.iterrows(): ws_ams.append([row["MATCH_STATUS"], int(row["Rreshta"]), int(row["Qty"]), float(row["Vlere"])])
    for r in range(2, ws_ams.max_row + 1): ws_ams.cell(row=r, column=4).number_format = "#,##0.00"
    add_table_safe(wb, ws_ams, "tblAPIMatchStats", f"A1:D{ws_ams.max_row}"); auto_fit(ws_ams)

    # Brand_Alias_Report
    try:
        ws_br = wb.create_sheet("Brand_Alias_Report")
        ws_br.append(["brand","winner_api","winner_count","total","candidates"]); style_header(ws_br, row=1)
        for _, rr in alias_report.iterrows():
            ws_br.append([rr.get("brand",""), rr.get("winner_api",""), int(rr.get("winner_count",0)),
                          int(rr.get("total",0)), rr.get("candidates","")])
        add_table_safe(wb, ws_br, "tblBrandAlias", f"A1:E{ws_br.max_row}")
        auto_fit(ws_br)
    except Exception as e:
        log(f"[WARN] Alias report sheet: {e}")

    # Watchlist
    ws_neg = wb.create_sheet("Watchlist")
    neg = df_review[df_review["Stanje"] < 0].copy().sort_values("Stanje")
    ws_neg.append(["Artikal","Furnitor","Stanje","Izlaz","Mes_Muj","Porosit_final","Sifra","Barkod"]); style_header(ws_neg, row=1)
    for _, row in neg.iterrows():
        ws_neg.append([row["Artikal"], row["Furnitor_best"], float(row["Stanje"]), float(row["Izlaz"]),
                       float(row["Mes_Muj"]), int(row["Porosit_final"]), row["Sifra_actual"], row["Barkod_actual"]])
    for r in range(2, ws_neg.max_row + 1): ws_neg.cell(row=r, column=3).number_format = "#,##0"
    add_table_safe(wb, ws_neg, "tblWatch", f"A1:H{ws_neg.max_row}"); auto_fit(ws_neg)

    # Top20
    ws_top = wb.create_sheet("Top20")
    top = df_review.sort_values("value", ascending=False).head(20)
    ws_top.append(["Artikal","Sifra","Furnitor","Kolicina","Vlera","Cmim","Stanje","Izlaz"]); style_header(ws_top, row=1)
    for _, row in top.iterrows():
        ws_top.append([row["Artikal"], row["Sifra_actual"], row["Furnitor_best"], int(row["Porosit_final"]),
                       float(row["value"]), float(row["Cmim_best"]), float(row["Stanje"]), float(row["Izlaz"])])
    for r in range(2, ws_top.max_row + 1):
        ws_top.cell(row=r, column=5).number_format = "#,##0.00"
        ws_top.cell(row=r, column=6).number_format = "#,##0.00"
    add_table_safe(wb, ws_top, "tblTop20", f"A1:H{ws_top.max_row}"); auto_fit(ws_top)

    # Exceptions
    ws_exc = wb.create_sheet("Exceptions")
    exc = df_review[
        (df_review["Match_method"] == "UNMATCHED") |
        ((df_review["Cmim_best"] <= 0) & (df_review["Porosit_final"] > 0)) |
        ((df_review["Target_Stock"] <= 0) & (df_review["Porosit_final"] > 0)) |
        ((df_review["Mes_Muj"] == 0) & (df_review["Izlaz"] > 0)) |
        ((df_review["Stanje"] < 0) & (df_review["Needs_Order"] == "Jo"))
    ].copy()
    ws_exc.append(["Artikal","Sifra","Furnitor","Match","Winner_Reason","Kolicina","Vlera","Cmim","Stanje","Min/Target","Mes_Muj","Izlaz","Ulaz","Barkod"])
    style_header(ws_exc, row=1)
    for _, row in exc.iterrows():
        ws_exc.append([row["Artikal"], row["Sifra_actual"], row["Furnitor_best"], row["Match_method"], row.get("Winner_Reason",""),
                       int(row["Porosit_final"]), float(row["value"]), float(row["Cmim_best"]), float(row["Stanje"]),
                       float(row["Target_Stock"]), float(row["Mes_Muj"]), float(row["Izlaz"]), float(row["Ulaz"]),
                       row["Barkod_actual"]])
    for r in range(2, ws_exc.max_row + 1):
        ws_exc.cell(row=r, column=7).number_format = "#,##0.00"
        ws_exc.cell(row=r, column=8).number_format = "#,##0.00"
    add_table_safe(wb, ws_exc, "tblExceptions", f"A1:N{ws_exc.max_row}"); auto_fit(ws_exc)

    # ABC
    try:
        total_val = df_review["value"].sum();  total_val = total_val if total_val>0 else 1.0
        def cls(p): return "A" if p<=0.80 else ("B" if p<=0.95 else "C")
        ws_abc_i = wb.create_sheet("ABC_Items")
        items = (df_review.groupby(["Sifra_actual","Artikal"], dropna=False)["value"].sum().reset_index()
                 .sort_values("value", ascending=False))
        items["Share"] = items["value"]/total_val; items["CumShare"]=items["Share"].cumsum(); items["Class"]=items["CumShare"].map(cls)
        ws_abc_i.append(["Sifra","Artikal","Value","Share","CumShare","Class"]); style_header(ws_abc_i, row=1)
        for _, r in items.iterrows():
            ws_abc_i.append([r["Sifra_actual"], r["Artikal"], float(r["value"]), float(r["Share"]), float(r["CumShare"]), r["Class"]])
        for rr in range(2, ws_abc_i.max_row + 1):
            ws_abc_i.cell(row=rr, column=3).number_format = "#,##0.00"
            ws_abc_i.cell(row=rr, column=4).number_format = "0.00%"
            ws_abc_i.cell(row=rr, column=5).number_format = "0.00%"
        add_table_safe(wb, ws_abc_i, "tblABCItems", f"A1:F{ws_abc_i.max_row}"); auto_fit(ws_abc_i)

        ws_abc_s = wb.create_sheet("ABC_Suppliers")
        sups2 = (df_review.groupby("Furnitor_best")["value"].sum().reset_index().sort_values("value", ascending=False))
        sups2["Share"] = sups2["value"]/total_val; sups2["CumShare"]=sups2["Share"].cumsum(); sups2["Class"]=sups2["CumShare"].map(cls)
        ws_abc_s.append(["Furnitor","Value","Share","CumShare","Class"]); style_header(ws_abc_s, row=1)
        for _, r in sups2.iterrows():
            ws_abc_s.append([r["Furnitor_best"], float(r["value"]), float(r["Share"]), float(r["CumShare"]), r["Class"]])
        for rr in range(2, ws_abc_s.max_row + 1):
            ws_abc_s.cell(row=rr, column=2).number_format = "#,##0.00"
            ws_abc_s.cell(row=rr, column=3).number_format = "0.00%"
            ws_abc_s.cell(row=rr, column=4).number_format = "0.00%"
        add_table_safe(wb, ws_abc_s, "tblABCSup", f"A1:E{ws_abc_s.max_row}"); auto_fit(ws_abc_s)
    except Exception: pass

    # ---------------- PO_<Furnitor> ----------------
    # Krijo indekset për ID të furnitorëve për CSV/PO
    SUP_BARK, SUP_SIG = {}, {}
    sub = (furn_tbl[furn_tbl["Barkod"]!=""]
           .sort_values(["Furnitor","Barkod","PriceValid","Price"])
           .groupby(["Furnitor","Barkod"], as_index=False).first())
    for sup, grp in sub.groupby("Furnitor"): SUP_BARK[sup] = dict(zip(grp["Barkod"], grp["Sifra"]))
    sub2 = (furn_tbl.sort_values(["Furnitor","Signature","PriceValid","Price"])
            .groupby(["Furnitor","Signature"], as_index=False).first())
    for sup, grp in sub2.groupby("Furnitor"): SUP_SIG[sup] = dict(zip(grp["Signature"], grp["Sifra"]))

    for furn in sorted([f for f in df_review["Furnitor_best"].dropna().unique() if str(f).strip() != ""]):
        sup_u = str(furn).upper().strip()
        block = df_review[(df_review["Furnitor_best"] == furn) & (df_review["Porosit_final"] > 0)].copy()
        sh = f"PO_{str(sup_u)[:25]}"
        ws_po = wb.create_sheet(sh)
        ws_po.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws_po["A1"] = f"TREBOVANJE PËR FURNITORIN: {sup_u} | {datetime.now():%Y-%m-%d %H:%M}"
        ws_po["A1"].font = Font(bold=True); ws_po["A1"].alignment = Alignment(horizontal="center")
        ws_po.append([]); ws_po.append(["Sifra/ID","Barkod","Artikal","Kolicina","Stanje","Izlaz","Ulaz"])
        style_header(ws_po, row=3)
        if not block.empty:
            block = block.sort_values(["Porosit_final", "Artikal"], ascending=[False, True])
            for _, row in block.iterrows():
                sup_id = sup_id_for_row(row, sup_u, SUP_BARK, SUP_SIG)  # pse: ID furnitori, fallback barkod
                ws_po.append([
                    sup_id,
                    row.get("Barkod_actual", ""),
                    row["Artikal"],
                    int(row["Porosit_final"]),
                    float(row["Stanje"]),
                    float(row["Izlaz"]),
                    float(row.get("Ulaz", 0.0)),
                ])
        last_po = ws_po.max_row
        if last_po >= 4:
            ws_po.conditional_formatting.add(f"D4:D{last_po}", DataBarRule(start_type="min", end_type="max", color="63BE7B"))
            ws_po.conditional_formatting.add(f"E4:E{last_po}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")))
            ws_po.freeze_panes = "A4"
            add_table_safe(wb, ws_po, f"tbl_{re.sub(r'[^A-Za-z0-9_]', '_', sh)}", f"A3:G{last_po}")
            total_row = last_po + 1
            ws_po.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
            ws_po.cell(row=total_row, column=4, value=f"=SUBTOTAL(9,D4:D{last_po})").font = Font(bold=True)
            ws_po.cell(row=total_row, column=4).number_format = "#,##0"
        auto_fit(ws_po)

    # ---------------- Quick Views ----------------
    try:
        q1 = df_review[df_review["Porosit_final"] > 0].copy()
        ws_q1 = wb.create_sheet("View_OrderOnly")
        cols = ["Furnitor_best","Sifra_actual","Artikal","Porosit_final","value","Cmim_best",
                "Rabat_best","Rabat_others_max","Rabat_others_avg",
                "Winner_Reason","Stanje","Izlaz","Ulaz","Match_method"]
        ws_q1.append(cols); style_header(ws_q1, row=1)
        for _, r in q1[cols].sort_values(["Furnitor_best","value"], ascending=[True,False]).iterrows():
            ws_q1.append([r[c] for c in cols])
        for rr in range(2, ws_q1.max_row + 1):
            ws_q1.cell(row=rr, column=4).number_format = "#,##0"
            ws_q1.cell(row=rr, column=5).number_format = "#,##0.00"
            ws_q1.cell(row=rr, column=6).number_format = "#,##0.00"
            ws_q1.cell(row=rr, column=7).number_format = "0.00%"
            ws_q1.cell(row=rr, column=8).number_format = "0.00%"
            ws_q1.cell(row=rr, column=9).number_format = "0.00%"
        add_table_safe(wb, ws_q1, "tblOrderOnly", f"A1:N{ws_q1.max_row}"); auto_fit(ws_q1)

        q2 = df_review[df_review["Match_method"] == "UNMATCHED"].copy()
        ws_q2 = wb.create_sheet("View_Unmatched")
        cols2 = ["Sifra_actual","Barkod_actual","Artikal","Stanje","Izlaz","Mes_Muj","Porosit_final"]
        ws_q2.append(cols2); style_header(ws_q2, row=1)
        for _, r in q2[cols2].iterrows():
            ws_q2.append([r[c] for c in cols2])
        add_table_safe(wb, ws_q2, "tblUnmatched", f"A1:G{ws_q2.max_row}"); auto_fit(ws_q2)

        q3 = df_review.sort_values("value", ascending=False).head(100)
        ws_q3 = wb.create_sheet("View_HighValue")
        cols3 = ["Furnitor_best","Sifra_actual","Artikal","Porosit_final","value","Cmim_best"]
        ws_q3.append(cols3); style_header(ws_q3, row=1)
        for _, r in q3[cols3].iterrows():
            ws_q3.append([r[c] for c in cols3])
        for rr in range(2, ws_q3.max_row + 1):
            ws_q3.cell(row=rr, column=5).number_format = "#,##0.00"
            ws_q3.cell(row=rr, column=6).number_format = "#,##0.00"
        add_table_safe(wb, ws_q3, "tblHighValue", f"A1:F{ws_q3.max_row}"); auto_fit(ws_q3)
    except Exception: pass

    # Legend
    ws_leg = wb.create_sheet("Legend")
    ws_leg["A1"] = "Legend — Shpjegime"; ws_leg["A1"].font = Font(size=14, bold=True)
    ws_leg.append(["Fusha","Vlerë","Shpjegim"]); style_header(ws_leg, row=2)
    legends = [
        ("Match_method","BARKOD","Match exakt me barkod (#1)"),
        ("Match_method","SIFRA","Match me Sifra Artikla (#2)"),
        ("Match_method","SIGNATURE","Emër i normalizuar (çmimi më i ulët)"),
        ("Match_method","API_DOSE","API+dozë(+formë), pa paketim"),
        ("Match_method","SIGNATURE2","Signature semantike pa paketim/prodhues"),
        ("Match_method","API_NEAR","API+unit+form të njëjta, dozë afër"),
        ("MATCH_STATUS","EXACT","API+dozë(+formë) dhe paketim i njëjtë"),
        ("MATCH_STATUS","FUZZY","API+dozë(+formë) e njëjtë; paketimi ndryshon"),
        ("MATCH_STATUS","NO_MATCH","Asnjë furnitor me API+dozë të njëjtë"),
        ("Winner_Reason","PRICE_MIN_*","Furnitori u zgjodh me çmimin efektiv më të ulët"),
        ("Rabat_best","%","Rabati i furnitorit fitues"),
        ("Rabat_others_max","%","Rabati maksimal te të tjerët për të njëjtin grup"),
        ("Rabat_others_avg","%","Rabati mesatar te të tjerët për të njëjtin grup"),
        ("Order Policy", f"TARGET_MODE={TARGET_MODE}", "Si llogaritet Target_Stock"),
        ("Order Policy", f"IGNORE_ULAZ_IN_ORDER={IGNORE_ULAZ_IN_ORDER}", "ULAZ injorohet në Available"),
    ]
    for f,v,s in legends: ws_leg.append([f,v,s])
    add_table_safe(wb, ws_leg, "tblLegend", f"A2:C{ws_leg.max_row}"); auto_fit(ws_leg)

    # ---------------- CSV EXPORT ----------------
    try:
        for c in ["Furnitor_best","Match_method","Sifra_actual","Barkod_actual","Artikal","Porosit_final"]:
            if c not in df_review.columns: df_review[c] = ""
        # indekset për CSV (ripërdor SUP_BARK/SUP_SIG nga sipër)
        def _export_for_supplier(df_rev: pd.DataFrame, sup_name: str):
            sup_u = str(sup_name).upper().strip()
            base = df_rev[df_rev["Porosit_final"] > 0].copy()
            mask_sup = base["Furnitor_best"].astype(str).str.upper().str.strip().eq(sup_u)
            blk = base[mask_sup].copy()
            if ONLY_DIGIT_NAMES:
                blk = blk[blk["Artikal"].astype(str).str.contains(r"\d", regex=True, na=False)]
            blk["Signature"] = blk["Artikal"].map(norm_name)
            sup_bark = SUP_BARK.get(sup_u, {}); sup_sig = SUP_SIG.get(sup_u, {})
            blk["ID_sup"] = blk["Barkod_actual"].map(lambda x: sup_bark.get(str(x).strip(), ""))
            needs_sig = blk["ID_sup"].eq("") | blk["ID_sup"].isna()
            blk.loc[needs_sig,"ID_sup"] = blk.loc[needs_sig,"Signature"].map(lambda s: sup_sig.get(s, ""))
            if STRICT_SIFRA_ONLY:
                blk["ID"] = blk["ID_sup"].fillna("").astype(str).str.strip()
            else:
                blk["ID"] = blk["ID_sup"].fillna("").astype(str).str.strip()
                miss = blk["ID"].eq("")
                blk.loc[miss,"ID"] = blk.loc[miss,"Barkod_actual"].astype(str).str.strip()
            blk = blk[blk["ID"] != ""]
            log(f"[CSV-DIAG] {sup_u}: rows={len(blk)}")
            if blk.empty:
                log(f"[CSV] {sup_u}: asnjë rresht për eksport."); return
            simple = blk[["ID","Porosit_final"]].rename(columns={"ID":"Sifra","Porosit_final":"Kolicina"})
            full = blk[["ID","Artikal","Porosit_final","Cmim_best","Stanje","Izlaz","Ulaz"]].rename(columns={"ID":"Sifra","Porosit_final":"Kolicina","Cmim_best":"Cmim"})
            if DECIMAL_COMMA:
                for col in ["Kolicina","Cmim","Stanje","Izlaz","Ulaz"]:
                    if col in full.columns: full[col] = full[col].map(lambda x: str(x).replace(".", ","))
                simple["Kolicina"] = simple["Kolicina"].map(lambda x: str(x).replace(".", ","))
            ts_csv = datetime.now().strftime("%Y%m%d_%H%M")
            csv1 = os.path.join(out_dir, f"{sup_u}_order_sifra_qty_{ts_csv}.csv")
            csv2 = os.path.join(out_dir, f"{sup_u}_order_sifra_full_{ts_csv}.csv")
            simple.to_csv(csv1, sep=CSV_DELIM, index=False, encoding="utf-8-sig", header=CSV_HEADER)
            full.to_csv(csv2, sep=CSV_DELIM, index=False, encoding="utf-8-sig", header=CSV_HEADER)
            log(f"[CSV] {sup_u} -> {csv1} | {csv2} (rows={len(blk)})")
        if EXPORT_ALL_SUPPLIERS:
            sups = sorted(df_review.loc[df_review["Porosit_final"] > 0, "Furnitor_best"].dropna().unique())
            for s in sups: _export_for_supplier(df_review, str(s))
        else:
            for s in SUPPLIERS: _export_for_supplier(df_review, s)
    except Exception as e:
        log(f"[WARN] CSV export: {e}")

    # ---------------- Save workbook ----------------
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    base = f"POROSIT_FULL_{ts}"
    out_path = os.path.join(out_dir, f"{base}.xlsx")
    v = 2
    while os.path.exists(out_path):
        out_path = os.path.join(out_dir, f"{base}_v{v}.xlsx")
        v += 1
    try:
        wb.save(out_path)
        log(f"[DONE] Saved: {out_path}")
    except PermissionError as e:
        alt = os.path.join(out_dir, f"{base}_UNLOCKME.xlsx")
        log(f"[WARN] File i hapur nga Excel? Po ruaj: {alt}  ({e})")
        wb.save(alt)
        log(f"[DONE] Saved (alt): {alt}")

if __name__ == "__main__":
    try:
        main()
    except Exception:
        print("[FATAL] Exception:\n" + traceback.format_exc(), flush=True)
        sys.exit(99)
import pandas as pd
import os

def calculate_reorder():
    file_path = os.path.join(os.path.dirname(__file__), '../data/inventory_data.csv')
    df = pd.read_csv(file_path)

    df["need_raw"] = df["min_stock"] - df["stock_now"]
    df["need_raw"] = df["need_raw"].apply(lambda x: x if x > 0 else 0)
    df["qty_to_order_sug"] = (df["need_raw"] / df["pack_size"]).round().fillna(0).astype(int)

    return df[[
        "sku_id", "sku_name", "supplier_id", "pack_size", 
        "stock_now", "min_stock", "need_raw", "qty_to_order_sug"
    ]]
