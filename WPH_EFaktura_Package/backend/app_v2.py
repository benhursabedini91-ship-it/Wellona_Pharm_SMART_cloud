SSH direkt nga PowerShell (më trego nëse ke lidhur)import os
import urllib.parse
import sys
import csv
import io
import datetime
import hashlib
import json
from decimal import Decimal, ROUND_HALF_UP
from functools import lru_cache
from flask import Flask, jsonify, request, send_from_directory, Response, make_response
import requests
from flask import send_file
from dotenv import load_dotenv
from db import fetch_all

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Faktura AI integration
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__)), 'app'))
try:
    from faktura_ai_mvp import parse_invoice_xml, load_lookup_from_db, load_config, write_csv, to_number, ensure_dir, dtstamp
    from faktura_import import parse_invoice_xml as parse_sopharma_xml, insert_kalkulacija, MP_CONFIG
    import psycopg2
    # Optional: eFaktura client
    try:
        from efaktura_client import fetch_to_staging
        EFAKTURA_AVAILABLE = True
    except Exception:
        EFAKTURA_AVAILABLE = False
    FAKTURA_AI_AVAILABLE = True
    SOPHARMA_ERP_AVAILABLE = True
except ImportError as e:
    FAKTURA_AI_AVAILABLE = False
    SOPHARMA_ERP_AVAILABLE = False
    print(f"Warning: Faktura AI modules not available: {e}")

# Monitoring integration
try:
    from monitoring import monitor, send_telegram_alert
    MONITORING_AVAILABLE = True
except ImportError as e:
    MONITORING_AVAILABLE = False
    print(f"Warning: Monitoring module not available: {e}")

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env"))

app = Flask(__name__)
# Public UI directory (serves orders_pro_plus.html)
UI_PUBLIC_PATH = os.path.join(os.path.dirname(__file__), 'public')

# Simple in-memory cache for /api/orders (TTL: 5 minutes)
_orders_cache = {}
_cache_ttl = 300  # seconds

# ============ POST /api/orders/{supplier} Setup ============
OUT_DIR = r"C:\Wellona\wphAI\out\orders"
os.makedirs(OUT_DIR, exist_ok=True)

# Feature flags
USE_DB = os.getenv("WPH_APP_USE_DB", "0") == "1"
PGAPP_DSN = os.getenv("PGAPP_DSN", "")

# Banned words filter (can be disabled by clearing env)
BANNED_WORDS_STR = os.getenv("BANNED_WORDS", "igla,igle,spric,rukavica,rukavice,contour plus,maske,maska")
BANNED_WORDS = [w.strip().lower() for w in BANNED_WORDS_STR.split(",") if w.strip()]

# Daily budget cap (0 = unlimited)
try:
    DAILY_BUDGET_RSD = Decimal(os.getenv("DAILY_BUDGET_RSD", "1000000"))
except:
    DAILY_BUDGET_RSD = Decimal("1000000")

def ceil_to_pack(q_need: Decimal, pack: Decimal) -> Decimal:
    """Round qty UP to nearest multiple of pack size"""
    if pack <= 0:
        pack = Decimal("1")
    units = (q_need / pack).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    if (units * pack) < q_need:
        units += 1
    return units * pack

def sanitize_items(items):
    """
    Normalize fields, filter banned words, apply pakovanje rounding, budget cap.
    Returns: (list of dicts, total_value)
    """
    rows = []
    total = Decimal("0")
    
    for it in items:
        # Banned words check
        name = (it.get("naziv") or "").lower()
        if any(b in name for b in BANNED_WORDS):
            continue
        
        # Extract and normalize
        qty = Decimal(str(it.get("qty", 0) or 0))
        pack = Decimal(str(it.get("pakovanje", 1) or 1))
        
        # Apply pakovanje rounding
        qty = ceil_to_pack(qty, pack)
        
        if qty <= 0:
            continue
        
        price = Decimal(str(it.get("unit_cost") or it.get("price") or 0))
        line = qty * price
        
        rows.append({
            "sifra": (it.get("sifra") or "").strip(),
            "barkod": (it.get("barkod") or "").strip(),
            "naziv": (it.get("naziv") or "").strip(),
            "qty": qty,
            "unit_cost": price,
            "line_total": line,
            "pakovanje": pack
        })
        total += line
    
    # Budget cap (greedy by line_total ascending = keep cheaper items)
    if DAILY_BUDGET_RSD > 0 and total > DAILY_BUDGET_RSD:
        rows.sort(key=lambda x: x["line_total"])
        kept, run = [], Decimal("0")
        for r in rows:
            if run + r["line_total"] <= DAILY_BUDGET_RSD:
                kept.append(r)
                run += r["line_total"]
        rows, total = kept, run
    
    return rows, total


def _safe_fetch_all(sql, params=None, default=None):
    """Helper to execute fetch_all safely in backend app.
    Returns default when USE_DB not enabled; otherwise attempts fetch_all and raises on error.
    """
    if not USE_DB:
        app.logger.debug("DB disabled (WPH_APP_USE_DB=0) - returning default for fetch_all")
        return default if default is not None else []
    try:
        if params is None:
            return fetch_all(sql)
        return fetch_all(sql, params)
    except Exception as e:
        app.logger.exception("DB query failed: %s", e)
        raise

# ============ Routes ============

@app.get("/health")
def health():
    return jsonify({
        "status": "OK",
        "db_host": os.getenv("WPH_DB_HOST","127.0.0.1"),
        "db_name": os.getenv("WPH_DB_NAME","wph_ai"),
        "app_port": int(os.getenv("APP_PORT","8055")),
        "cache_size": len(_orders_cache),
        "cache_ttl_sec": _cache_ttl
    })


# Serve swagger.json for local docs UI
@app.route('/api/docs/swagger.json')
def api_swagger_json():
    try:
        # Prefer project swagger if present
        path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'app', 'Faktura AI', 'swagger.json')
        if not os.path.exists(path):
            # fallback to root-level swagger
            path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', 'app', 'Faktura AI', 'swagger.json')
        with open(path, 'r', encoding='utf-8') as fh:
            data = fh.read()
        return Response(data, mimetype='application/json')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/ui/swagger')
def ui_swagger():
    # Serve the static swagger UI page we added to public/
    swagger_file = os.path.join(os.path.dirname(__file__), 'public', 'swagger_ui.html')
    return send_file(swagger_file)

@app.post("/api/cache/clear")
def clear_cache():
    """Clear all cached orders data"""
    global _orders_cache
    cache_size = len(_orders_cache)
    _orders_cache.clear()
    return jsonify({
        "status": "cleared",
        "entries_removed": cache_size,
        "timestamp": datetime.datetime.now().isoformat()
    })

@app.get("/api/suppliers")
def api_suppliers():
    """Get list of active suppliers"""
    try:
        rows = _safe_fetch_all("SELECT supplier_id, code, name FROM ref.suppliers WHERE is_active = true ORDER BY name", None, default=[])
        return jsonify(rows)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        app.logger.exception("Failed to fetch suppliers")
        return jsonify({"error": "Database unavailable or query failed", "message": str(e), "traceback": tb}), 503

@app.get("/")
def home():
    return jsonify({
        "message": "WPH AI Order Brain API v2",
        "primary_endpoints": [
            "/api/orders/v2 - 🚀 PRIMARY: Fast snapshot-based orders (use this!)",
            "/health - Health check",
            "/api/suppliers - List active suppliers"
        ],
        "legacy_endpoints": [
            "/api/orders - ⚠️ LEGACY: Slow function-based (deprecated)",
            "/api/orders/phoenix - Phoenix orders",
            "/api/orders/<supplier> [POST] - Submit order"
        ],
        "ui": [
            "/ui - Web UI",
            "/smart - Smart orders"
        ]
    })

def get_sales_mv(sales_window: int) -> str:
    """
    Map sales_window to MV:
      7   -> ops._sales_7d
      15  -> ops._sales_15d
      30  -> ops._sales_30d
      60  -> ops._sales_60d
      180 -> ops._sales_180d
    """
    if sales_window <= 10:
        return "ops._sales_7d"
    elif sales_window <= 20:
        return "ops._sales_15d"
    elif sales_window <= 45:
        return "ops._sales_30d"
    elif sales_window <= 120:
        return "ops._sales_60d"
    else:
        return "ops._sales_180d"  # 180+ days

@app.get("/api/orders/v2")
def api_orders_v2():
    """
    🚀 PRIMARY API ENDPOINT - Use this for all future work!
    
    Uses ops.orders_30d_snapshot materialized view with:
    - MIN_ZALIHA policy (ref.min_zaliha_policy_v2)
    - FLOW detection (0_FLOW, LOW_FLOW, NORMAL_FLOW, GOOD_FLOW, HIGH_FLOW)
    - Smart qty calculation (LEAST of 30d vs 180d/6)
    - Alarm status (stock <= min_zaliha)
    - SUPPLIER DATA via BARKOD JOIN (cheapest_supplier)
    
    Benefits:
    - Fast: Pre-computed snapshot (5380 products in ~50ms)
    - Complete: 77% products have supplier/price data
    - Consistent: Single source of truth for order calculations
    
    Parameters:
    - q: Search query (emri, sifra, barkod)
    - include_zero: 1=show 0_FLOW products, 0=hide (default 0)
    - supplier: Filter by supplier (PHOENIX, VEGA, SOPHARMA, ZEGIN)
    - download: csv|xlsx for download format
    """
    # Parse parameters
    search_query = request.args.get("q", "").strip() or None
    include_zero = request.args.get("include_zero", "0") == "1"
    suppliers = request.args.getlist("supplier")
    suppliers = [s.strip().upper() for s in suppliers if s.strip()]
    
    # Base query from snapshot
    sql = """
        SELECT 
            sifra, emri, barkod, current_stock, 
            monthly_sales, avg_daily_sales, days_cover,
            min_zaliha, qty_to_order, alarm_active, flow_status,
            supplier_name, base_price, rabat_pct, effective_price,
            sales_180d, last_sale_date
        FROM ops.orders_30d_snapshot
        WHERE 1=1
    """
    params = []
    
    # Filter: include_zero (exclude 0_FLOW if False)
    if not include_zero:
        sql += " AND flow_status != '0_FLOW'"
    
    # Filter: search query
    if search_query:
        sql += " AND (LOWER(emri) LIKE %s OR LOWER(sifra) LIKE %s OR LOWER(barkod) LIKE %s)"
        pattern = f"%{search_query.lower()}%"
        params.extend([pattern, pattern, pattern])
    
    # Filter: suppliers
    if suppliers:
        placeholders = ','.join(['%s'] * len(suppliers))
        sql += f" AND UPPER(supplier_name) IN ({placeholders})"
        params.extend(suppliers)
    
    # Order by priority (alarm first, then flow, then qty)
    sql += " ORDER BY alarm_active DESC, flow_status DESC, qty_to_order DESC"
    
    try:
        rows = _safe_fetch_all(sql, params, default=[])
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        app.logger.exception("api_orders_v2 DB query failed")
        return jsonify({"error": "Database error: unable to fetch orders", "message": str(e), "traceback": tb}), 503
    
    # Format decimals
    for row in rows:
        if row.get('avg_daily_sales') is not None:
            row['avg_daily_sales'] = round(float(row['avg_daily_sales']), 2)
        if row.get('days_cover') is not None:
            row['days_cover'] = round(float(row['days_cover']), 1)
        if row.get('current_stock') is not None:
            row['current_stock'] = round(float(row['current_stock']), 0)
        if row.get('monthly_sales') is not None:
            row['monthly_sales'] = round(float(row['monthly_sales']), 1)
        if row.get('sales_180d') is not None:
            row['sales_180d'] = round(float(row['sales_180d']), 1)
        if row.get('min_zaliha') is not None:
            row['min_zaliha'] = round(float(row['min_zaliha']), 0)
        if row.get('qty_to_order') is not None:
            row['qty_to_order'] = round(float(row['qty_to_order']), 0)
        if row.get('base_price') is not None:
            row['base_price'] = round(float(row['base_price']), 2)
        if row.get('rabat_pct') is not None:
            row['rabat_pct'] = round(float(row['rabat_pct']), 2)
        if row.get('effective_price') is not None:
            row['effective_price'] = round(float(row['effective_price']), 2)
    
    # Check if download format requested
    download = request.args.get("download", "").lower()
    if download == "csv":
        return _csv_response(rows)
    elif download == "xlsx":
        return _xlsx_response(rows)
    
    return jsonify(rows)


@app.get("/api/orders")
def api_orders():
    """
    ⚠️  LEGACY ENDPOINT - Use /api/orders/v2 instead!
    
    This endpoint uses wph_core.get_orders_v4_with_rabat() function which:
    - Slower: Computes on-the-fly (2-3 seconds)
    - Incomplete: Missing supplier data for many products
    - Deprecated: Will be removed in future versions
    
    Kept only for backward compatibility with old scripts.
    """
    # Parse parameters
    try:
        target_days = int(request.args.get("target_days", 28))
    except:
        target_days = 28
    
    # Optional legacy param (kept for backward-compat UI chips); backend now prefers explicit date range
    sales_window = request.args.get("sales_window")
    
    # New unified scope parameter: product_scope=sales|all
    #  sales -> only products with sales (maps to include_zero=0)
    #  all   -> include all products (maps to include_zero=1)
    # Backward compatibility: if include_zero explicitly provided, it overrides product_scope.
    product_scope = (request.args.get("product_scope") or "").strip().lower()
    include_zero_raw = request.args.get("include_zero")
    if include_zero_raw is not None:
        include_zero = (include_zero_raw == "1")
    else:
        include_zero = product_scope in ("all", "all_products", "allitems", "allproduct")
    search_query = request.args.get("q", "").strip() or None
    
    # New: explicit date range (let DB defaults kick in when missing)
    date_from = request.args.get("date_from") or None
    date_to = request.args.get("date_to") or None
    
    # Get supplier filter (multiple values supported)
    suppliers = request.args.getlist("supplier")
    # Filter empty strings
    suppliers = [s.strip().upper() for s in suppliers if s.strip()]
    
    # ✅ FIX: Always pass None to let SQL function auto-select cheapest supplier (ERP mode)
    # If you want supplier filtering, uncomment the next line and comment the one after
    # supplier_param = suppliers if suppliers else None
    supplier_param = None  # Force ERP mode: auto-select cheapest supplier
    
    # Call DB function. Prefer 6-param signature (target_days,date_from,date_to,include_zero,search_query,suppliers)
    # but gracefully fallback to 5-param legacy (target_days,sales_window,include_zero,search_query,suppliers)
    rows = []
    legacy_error = None
    try:
        rows = _safe_fetch_all(
            "SELECT * FROM wph_core.get_orders(%s, %s, %s, %s, %s, %s)",
            [target_days, date_from, date_to, include_zero, search_query, supplier_param],
            default=[]
        )
    except Exception as e6:
        legacy_error = e6  # store for diagnostics if fallback also fails
        app.logger.warning("6-param get_orders failed, attempting 5-param fallback: %s", e6)
        try:
            # Fallback: interpret target_days as p_target_days and sales_window from target_days or default 30
            fallback_sales_window = target_days if target_days in (7,15,30,60,180) else 30
            rows = _safe_fetch_all(
                "SELECT * FROM wph_core.get_orders(%s, %s, %s, %s, %s)",
                [target_days, fallback_sales_window, include_zero, search_query, supplier_param],
                default=[]
            )
        except Exception as e5:
            import traceback
            tb = traceback.format_exc()
            app.logger.exception("api_orders legacy DB query failed (both signatures)")
            return jsonify({
                "error": "Database error: unable to fetch orders (legacy)",
                "message": str(e5),
                "traceback": tb,
                "attempted_signatures": {
                    "6_param_error": str(legacy_error),
                    "5_param_error": str(e5)
                }
            }), 503
    
    # 🔧 FIX: Format decimal numbers to avoid 80.000000000000000 display
    for row in rows:
        # Convert Decimal to float with proper precision
        if row.get('avg_daily_sales') is not None:
            row['avg_daily_sales'] = round(float(row['avg_daily_sales']), 2)
        if row.get('days_cover') is not None:
            row['days_cover'] = round(float(row['days_cover']), 1)
        if row.get('current_stock') is not None:
            row['current_stock'] = round(float(row['current_stock']), 0)
        if row.get('min_zaliha') is not None:
            row['min_zaliha'] = round(float(row['min_zaliha']), 0)
        if row.get('qty_to_order') is not None:
            row['qty_to_order'] = round(float(row['qty_to_order']), 0)
        # New rabat columns
        if row.get('base_price') is not None:
            row['base_price'] = round(float(row['base_price']), 2)
        if row.get('rabat_pct') is not None:
            row['rabat_pct'] = round(float(row['rabat_pct']), 2)
        if row.get('effective_price') is not None:
            row['effective_price'] = round(float(row['effective_price']), 2)
    
    # Check if download format requested
    download = request.args.get("download", "").lower()
    if download == "csv":
        return _csv_response(rows)
    elif download == "xlsx":
        return _xlsx_response(rows)
    
    return jsonify(rows)


def _csv_response(rows):
    """Generate CSV download response"""
    if not rows:
        return Response("", mimetype="text/csv")
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys(), delimiter=";")
    writer.writeheader()
    writer.writerows(rows)
    
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"orders_{ts}.csv"
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "text/csv; charset=utf-8-sig"
    return response


def _xlsx_response(rows):
    """Generate XLSX download response with styling"""
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 500
    
    if not rows:
        return Response("", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Header styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            # Convert Decimal to float for Excel
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"orders_{ts}.xlsx"
    
    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@app.get("/api/health/db")
def health_db():
    """Database health check endpoint"""
    try:
        rows = _safe_fetch_all("SELECT 1 AS ok", None, default=None)
        if rows and rows[0].get("ok") == 1:
            return jsonify({
                "status": "healthy",
                "database": "connected",
                "timestamp": datetime.datetime.now().isoformat()
            })
        else:
            return jsonify({
                "status": "unhealthy",
                "database": "query failed",
                "timestamp": datetime.datetime.now().isoformat()
            }), 500
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        app.logger.exception("Database health check failed")
        return jsonify({
            "status": "unhealthy",
            "database": "connection failed",
            "error": str(e),
            "traceback": tb,
            "timestamp": datetime.datetime.now().isoformat()
        }), 503


@app.post("/api/orders/<supplier>")
def post_order_supplier(supplier):
    """
    Create order for supplier. Always generates CSV if approved=true.
    Optionally persists to DB if WPH_APP_USE_DB=1 and PGAPP_DSN set.
    
    Payload example:
    {
      "approved": true,
      "approved_by": "Beni",
      "target_days": 28,
      "sales_window": 60,
      "items": [
        {"sifra":"000123","qty":11,"unit_cost":199.99,"naziv":"PROVE","barkod":"860..","pakovanje":6},
        {"sifra":"000124","qty":4,"unit_cost":249.99,"naziv":"PROVE 2","barkod":"860.."}
      ],
      "note": "opsionale"
    }
    """
    try:
        data = request.get_json(force=True)
    except:
        return jsonify({"error": "Invalid JSON payload"}), 400
    
    # NEW: Support UI payload format with edits + filters
    if "edits" in data and "filters" in data:
        # Payload from new Orders Pro+ UI
        approved_by = (data.get("approved_by") or "system").strip()
        filters = data.get("filters", {})
        edits = data.get("edits", {})
        
        # Fetch current orders from DB using filters
        target_days = int(filters.get("target_days", 28))
        sales_window = int(filters.get("sales_window", 30))
        scope = (filters.get("product_scope") or "").strip().lower()
        if "include_zero" in filters:  # legacy key wins if present
            include_zero = str(filters.get("include_zero")) == "1"
        else:
            include_zero = scope in ("all", "all_products", "allitems", "allproduct")
        search_query = filters.get("q") or None
        suppliers = filters.get("supplier", [])
        
        # Get orders from DB - always use 5-param version
        try:
            rows_db = _safe_fetch_all(
                "SELECT * FROM wph_core.get_orders(%s, %s, %s, %s, %s)",
                [target_days, sales_window, include_zero, search_query, suppliers if suppliers else None],
                default=[]
            )
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            app.logger.exception("post_order_supplier DB fetch failed")
            return jsonify({"error":"Database error while fetching orders for UI payload", "message": str(e), "traceback": tb}), 503
        
        # Apply edits to quantities
        items_raw = []
        for row in rows_db:
            sifra = row.get("sifra")
            # Check if user edited this row
            if sifra in edits and "qty" in edits[sifra]:
                qty = int(edits[sifra]["qty"])
            else:
                qty = int(row.get("qty_to_order") or 0)
            
            if qty > 0:
                items_raw.append({
                    "sifra": sifra,
                    "barkod": row.get("barkod") or "",
                    "naziv": row.get("emri") or "",
                    "qty": qty,
                    "unit_cost": 0,  # Will be filled by sanitize_items if available
                    "pakovanje": int(row.get("pack_size") or 1)
                })
        
        approved = True
        note = f"Generated from Orders Pro+ UI by {approved_by}"
        # Set for DB persistence
        sales_win = sales_window
    else:
        # OLD: Original payload format with items[] array
        items_raw = data.get("items") or []
        approved = bool(data.get("approved", False))
        approved_by = (data.get("approved_by") or "system").strip()
        target_days = int(data.get("target_days", 6))
        sales_win = int(data.get("sales_window", 60))
        note = data.get("note")
    
    # Sanitize, filter, round
    rows, total = sanitize_items(items_raw)
    
    if not rows:
        return jsonify({"error": "No valid items after filters/rounding"}), 400
    
    status = "approved" if approved else "draft"
    csv_path = None
    order_id = None
    
    # Generate CSV when approved
    if approved:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        daydir = os.path.join(OUT_DIR, datetime.date.today().isoformat())
        os.makedirs(daydir, exist_ok=True)
        fname = f"ORDER_{supplier.upper()}_{ts}.csv"
        csv_path = os.path.join(daydir, fname)
        
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=';')
            w.writerow(["sifra", "barkod", "naziv", "qty", "unit_cost", "line_total"])
            for r in rows:
                w.writerow([
                    r["sifra"],
                    r["barkod"],
                    r["naziv"],
                    str(r["qty"]),
                    str(r["unit_cost"]),
                    str(r["line_total"])
                ])
    
    # Optional DB persistence (only if flag + DSN present)
    if USE_DB and PGAPP_DSN:
        try:
            import psycopg2
            import psycopg2.extras
            conn = psycopg2.connect(PGAPP_DSN)
            try:
                with conn:
                    with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                        cur.execute("""
                            INSERT INTO app.orders (
                                supplier, status, target_days, sales_window,
                                created_by, approved_by, approved_at,
                                csv_path, note, meta
                            )
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            RETURNING id
                        """, [
                            supplier, status, target_days, sales_win, "webapp",
                            (approved_by if approved else None),
                            (datetime.datetime.utcnow() if approved else None),
                            csv_path, note,
                            psycopg2.extras.Json({"items": len(rows)})
                        ])
                        order_id = cur.fetchone()[0]
                        
                        for r in rows:
                            cur.execute("""
                                INSERT INTO app.order_items (
                                    order_id, sifra, barkod, naziv,
                                    qty, unit_cost, meta
                                )
                                VALUES (%s,%s,%s,%s,%s,%s,%s)
                            """, [
                                order_id, r["sifra"], r["barkod"], r["naziv"],
                                str(r["qty"]), str(r["unit_cost"]),
                                psycopg2.extras.Json({"pakovanje": str(r["pakovanje"])})
                            ])
            finally:
                conn.close()
        except Exception as e:
            # DB insert failed but CSV exists - log and continue
            print(f"[WARN] DB insert failed: {e}")
    
    return jsonify({
        "status": status,
        "supplier": supplier,
        "items": len(rows),
        "total_rsd": float(total),
        "csv": csv_path,
        "order_id": order_id
    }), 201

@app.get("/api/orders/phoenix")
def api_phoenix():
    try:
        target_days = float(request.args.get("target_days", 28))
    except Exception:
        target_days = 28.0
    target_days = max(1.0, min(100.0, target_days))

    try:
        sales_window = int(request.args.get("sales_window", 60))
    except Exception:
        sales_window = 60
    
    sales_mv = get_sales_mv(sales_window)

    try:
        rows = _safe_fetch_all(f"""
                WITH demand AS (
                    SELECT s.sifra, COALESCE(s.avg_daily,0) AS avg_daily, s.last_sale_date,
                                 (CURRENT_DATE - s.last_sale_date) <= 90 AS has_recent_sales,
                                 COALESCE(s.avg_daily*30,0) AS monthly_units
                    FROM {sales_mv} s
                ),
                inventory AS (SELECT sifra, qty AS stock FROM stg.stock_on_hand),
                calc AS (
                    SELECT COALESCE(d.sifra, i.sifra) AS sifra,
                                 COALESCE(i.stock,0) AS current_stock,
                                 COALESCE(d.avg_daily,0) AS avg_daily_sales,
                                 COALESCE(d.has_recent_sales,FALSE) AS has_recent_sales,
                                 (
                                     SELECT p.min_zaliha FROM ref.min_zaliha_policy_v2 p
                                     WHERE COALESCE(d.avg_daily,0)*30 >= p.range_from
                                         AND (p.range_to IS NULL OR COALESCE(d.avg_daily,0)*30 <= p.range_to)
                                     ORDER BY p.range_from DESC LIMIT 1
                                 ) AS min_zaliha
                    FROM demand d FULL OUTER JOIN inventory i ON d.sifra = i.sifra
                )
                SELECT ar.barkod,
                             CASE 
                                 WHEN NOT c.has_recent_sales THEN 
                                     CASE WHEN c.current_stock=0 AND c.avg_daily_sales>0 THEN 2 ELSE 0 END
                                 ELSE GREATEST(0, GREATEST(COALESCE(c.min_zaliha,0), CEIL(%s * c.avg_daily_sales)) - c.current_stock)
                             END AS qty
                FROM calc c
                JOIN eb_fdw.artikli ar ON c.sifra = ar.sifra
                WHERE ar.barkod IS NOT NULL AND ar.barkod <> ''''
                ORDER BY qty DESC
                """, [target_days], default=[])
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        app.logger.exception("api_phoenix DB query failed")
        return jsonify({"error": "Database error: unable to compute phoenix orders", "message": str(e), "traceback": tb}), 503
    csv_lines = ["barkod,qty"]
    for r in rows:
        qty = int(r.get("qty") or 0)
        if qty > 0:
            csv_lines.append(f"{r.get('barkod')},{qty}")
    return "\n".join(csv_lines), 200, {"Content-Type": "text/csv; charset=utf-8"}

@app.get("/orders/download")
def orders_download():
    try:
        latest=None
        latest_mtime=0.0
        for root, dirs, files in os.walk(OUT_DIR):
            for fn in files:
                if fn.lower().endswith(".csv"):
                    pth=os.path.join(root, fn)
                    m=os.path.getmtime(pth)
                    if m>latest_mtime:
                        latest_mtime=m
                        latest=pth
        if not latest:
            return jsonify({"error":"no csv found"}),404
        return send_from_directory(os.path.dirname(latest), os.path.basename(latest), as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}),500


@app.errorhandler(500)
def handle_500(e):
    # Graceful fallback: log and return a structured JSON error for API requests
    import traceback
    try:
        tb = traceback.format_exc()
    except Exception:
        tb = None

    app.logger.exception("Unhandled 500 error: %s", e)
    # If the request is for an API path, return JSON; otherwise return minimal text
    try:
        if request.path.startswith('/api'):
            return jsonify({
                'error': 'internal_server_error',
                'message': str(e),
                'traceback': tb
            }), 500
        else:
            return Response("Internal server error", status=500, mimetype='text/plain')
    except Exception:
        return jsonify({'error': 'internal error'}), 500


# ============ UI Routes ============
@app.route("/")
@app.route("/ui")
@app.route("/ui/")
@app.route("/ui/pro")
@app.route("/ui/pro/")
def ui_index():
    """🚀 PRIMARY UI - Orders Pro+ with snapshot API"""
    response = make_response(send_from_directory(os.path.join(os.path.dirname(__file__), "public"), "orders_pro_plus.html"))
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.get("/api/health")
def api_health():
    """Lightweight health for browser/curl checks"""
    try:
        return jsonify({
            "status": "ok",
            "ts": datetime.datetime.utcnow().isoformat() + "Z",
            "endpoints": [
                "/api/orders/v2",
                "/api/faktura/upload",
                "/api/faktura/list",
                "/api/health/db",
                "/api/health/full"
            ]
        })
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500

@app.route("/api/health/full", methods=["GET"])
def api_health_full():
    """
    🏥 Comprehensive health check with system metrics
    Includes: DB, CPU, Memory, Disk, Uptime
    """
    if not MONITORING_AVAILABLE:
        return jsonify({
            "status": "degraded",
            "error": "Monitoring module not available - install psutil"
        }), 503
    
    try:
        health_data = monitor.get_full_health()
        
        # Check for critical issues
        issues = monitor.check_critical_issues()
        if issues:
            health_data["issues"] = issues
        
        status_code = 200 if health_data["overall_status"] == "healthy" else 503
        return jsonify(health_data), status_code
    except Exception as e:
        import traceback
        return jsonify({
            "status": "error",
            "error": str(e),
            "traceback": traceback.format_exc()
        }), 500

@app.route("/api/monitoring/alert", methods=["POST"])
def api_monitoring_alert():
    """
    📢 Manually trigger a test alert (Telegram)
    POST body: {"message": "Test alert", "severity": "warning"}
    """
    if not MONITORING_AVAILABLE:
        return jsonify({"error": "Monitoring not available"}), 503
    
    try:
        data = request.get_json() or {}
        message = data.get("message", "Test alert from WPH Pharmacy")
        severity = data.get("severity", "warning")
        
        success = send_telegram_alert(message, severity)
        
        return jsonify({
            "success": success,
            "message": "Alert sent successfully" if success else "Alert failed (check Telegram config)"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============ FAKTURA AI ENDPOINTS ============
@app.route("/api/faktura/upload", methods=["POST"])
def api_faktura_upload():
    """
    📄 Upload and parse invoice XML/PDF
    
    Accepts: multipart/form-data with file
    Returns: {header, items, validation}
    """
    if not FAKTURA_AI_AVAILABLE:
        return jsonify({"error": "Faktura AI module not available"}), 503
    
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "Empty filename"}), 400
        
        # Save to temp location
        base = os.path.dirname(os.path.dirname(__file__))
        temp_dir = ensure_dir(os.path.join(base, "staging", "faktura_uploads"))
        ts = dtstamp()
        temp_path = os.path.join(temp_dir, f"{ts}_{file.filename}")
        file.save(temp_path)
        
        # Parse XML
        if not temp_path.lower().endswith('.xml'):
            return jsonify({"error": "Only XML invoices supported currently"}), 400
        
        hdr, items = parse_invoice_xml(temp_path)
        
        # Load validation lookup
        cfg_path = os.path.join(base, 'configs', 'faktura_ai.json')
        cfg = load_config(cfg_path)
        sifra_set, barcode_set = load_lookup_from_db(cfg['db'])
        
        # Validate items
        matched = 0
        for item in items:
            s = (item.get('sifra') or '').strip()
            b = (item.get('barcode') or '').strip()
            if (s and s in sifra_set) or (b and b in barcode_set):
                matched += 1
                item['valid'] = True
            else:
                item['valid'] = False
        
        match_rate = (matched / len(items) * 100) if items else 0.0
        
        # Calculate totals
        total_calc = sum(
            to_number(it.get('qty'), 0) * 
            to_number(it.get('price'), 0) * 
            (1 - to_number(it.get('rabat_pct'), 0) / 100.0)
            for it in items
        )
        total_hdr = to_number(hdr.get('total_amount'), None)
        
        tol_pct = float(cfg.get('tolerance', {}).get('total_pct', 0.1))
        total_ok = True
        if total_hdr is not None:
            delta_pct = abs(total_calc - total_hdr) / total_hdr * 100 if total_hdr else 0
            if delta_pct > tol_pct:
                total_ok = False
        
        status = 'CLEAN' if (match_rate >= 99.0 and total_ok) else 'NEEDS_REVIEW'
        
        # Save to output
        out_root = ensure_dir(cfg['outdir'])
        out_day = ensure_dir(os.path.join(out_root, datetime.datetime.now().strftime('%Y%m%d')))
        inv_id = (hdr.get('invoice_no') or os.path.splitext(file.filename)[0]).replace('/', '_')
        prefix = os.path.join(out_day, f'{inv_id}')
        
        write_csv(prefix + '_header.csv', [hdr], ['supplier', 'invoice_no', 'invoice_date', 'currency', 'total_amount'])
        write_csv(prefix + '_items.csv', items, ['sifra', 'barcode', 'name', 'qty', 'price', 'rabat_pct'])
        
        return jsonify({
            "success": True,
            "header": hdr,
            "items": items,
            "validation": {
                "total_items": len(items),
                "matched": matched,
                "match_rate_pct": round(match_rate, 2),
                "total_calc": round(total_calc, 2),
                "total_header": total_hdr,
                "total_ok": total_ok,
                "status": status
            },
            "output": {
                "header_csv": prefix + '_header.csv',
                "items_csv": prefix + '_items.csv'
            }
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/faktura/efaktura-pull", methods=["POST"])
def api_faktura_efaktura_pull():
    """
    ⬇️ Pull invoices from eFaktura API into staging (requires env WPH_EFAKT_API_KEY and URLs).

    Body (optional): {"days": 7} or {"from": "YYYY-MM-DD", "to": "YYYY-MM-DD"}
    """
    try:
        if not EFAKTURA_AVAILABLE:
            return jsonify({"error": "efaktura_client not available"}), 500
        data = request.get_json(silent=True) or {}
        if 'from' in data and 'to' in data:
            dfrom = datetime.datetime.fromisoformat(data['from']).date()
            dto = datetime.datetime.fromisoformat(data['to']).date()
        else:
            days = int(data.get('days', 7))
            dto = datetime.date.today()
            dfrom = dto - datetime.timedelta(days=days)

        base = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'staging', 'faktura_uploads')
        ensure_dir(base)
        count, paths = fetch_to_staging(dfrom, dto, base)
        return jsonify({"fetched": count, "paths": paths})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/faktura/list", methods=["GET"])
def api_faktura_list():
    """
    📋 List processed invoices
    
    Query params:
      - date: YYYYMMDD (default: today)
    
    Returns: [{invoice_no, supplier, items, status, files}, ...]
    """
    if not FAKTURA_AI_AVAILABLE:
        return jsonify({"error": "Faktura AI module not available"}), 503
    
    try:
        base = os.path.dirname(os.path.dirname(__file__))
        cfg_path = os.path.join(base, 'configs', 'faktura_ai.json')
        cfg = load_config(cfg_path)
        out_root = cfg['outdir']
        
        date_str = request.args.get('date', datetime.datetime.now().strftime('%Y%m%d'))
        out_day = os.path.join(out_root, date_str)
        
        if not os.path.isdir(out_day):
            return jsonify({"invoices": []})
        
        # Group files by invoice_no prefix
        invoices = {}
        for fn in os.listdir(out_day):
            if fn.endswith('_header.csv'):
                inv_id = fn.replace('_header.csv', '')
                invoices[inv_id] = {
                    'invoice_no': inv_id,
                    'files': {
                        'header': os.path.join(out_day, fn),
                        'items': os.path.join(out_day, f'{inv_id}_items.csv')
                    }
                }
                
                # Read header to get supplier
                try:
                    with open(os.path.join(out_day, fn), 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        row = next(reader, {})
                        invoices[inv_id]['supplier'] = row.get('supplier', '')
                        invoices[inv_id]['invoice_date'] = row.get('invoice_date', '')
                except Exception:
                    pass
                
                # Count items
                try:
                    with open(os.path.join(out_day, f'{inv_id}_items.csv'), 'r', encoding='utf-8') as f:
                        invoices[inv_id]['items'] = sum(1 for _ in f) - 1  # exclude header
                except Exception:
                    invoices[inv_id]['items'] = 0
        
        return jsonify({"invoices": list(invoices.values())})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/faktura/download/<date>/<invoice_no>/<file_type>", methods=["GET"])
def api_faktura_download(date, invoice_no, file_type):
    """
    💾 Download invoice CSV (header or items)
    
    Path params:
      - date: YYYYMMDD
      - invoice_no: Invoice identifier
      - file_type: 'header' or 'items'
    """
    try:
        base = os.path.dirname(os.path.dirname(__file__))
        cfg_path = os.path.join(base, 'configs', 'faktura_ai.json')
        cfg = load_config(cfg_path)
        out_day = os.path.join(cfg['outdir'], date)
        
        if file_type not in ['header', 'items']:
            return jsonify({"error": "Invalid file_type"}), 400
        
        filename = f"{invoice_no}_{file_type}.csv"
        filepath = os.path.join(out_day, filename)
        
        if not os.path.exists(filepath):
            return jsonify({"error": "File not found"}), 404
        
        return send_from_directory(out_day, filename, as_attachment=True)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/faktura/pending", methods=["GET"])
def api_faktura_pending():
    """
    📋 List pending invoices (uploaded but not imported to DB)
    
    Returns:
      {
        "pending": [
          {
            "filename": "invoice_20241114.xml",
            "path": "C:\\...\\staging\\faktura_uploads\\invoice_20241114.xml",
            "uploaded_at": "2024-11-14T10:30:00",
            "size_kb": 12.5,
            "supplier": "PHOENIX PHARMA",
            "invoice_no": "INV-2024-001",
            "items_count": 5
          }
        ]
      }
    """
    try:
        base = os.path.dirname(os.path.dirname(__file__))
        staging_dir = os.path.join(base, 'staging', 'faktura_uploads')
        ensure_dir(staging_dir)
        
        pending = []
        for fname in os.listdir(staging_dir):
            if not fname.lower().endswith('.xml'):
                continue
            
            fpath = os.path.join(staging_dir, fname)
            stat = os.stat(fpath)
            
            # Quick parse to get header info
            try:
                header, items = parse_invoice_xml(fpath)
                pending.append({
                    "filename": fname,
                    "path": fpath,
                    "uploaded_at": datetime.datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "size_kb": round(stat.st_size / 1024, 2),
                    "supplier": header.get('supplier', 'N/A'),
                    "invoice_no": header.get('invoice_no', 'N/A'),
                    "invoice_date": header.get('invoice_date', 'N/A'),
                    "items_count": len(items)
                })
            except Exception as e:
                # If parse fails, still show file
                pending.append({
                    "filename": fname,
                    "path": fpath,
                    "uploaded_at": datetime.datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "size_kb": round(stat.st_size / 1024, 2),
                    "supplier": "PARSE_ERROR",
                    "invoice_no": "N/A",
                    "items_count": 0,
                    "error": str(e)
                })
        
        return jsonify({"pending": pending})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/faktura/preview", methods=["POST"])
def api_faktura_preview():
    """
    🔍 Preview invoice before import (parse + match + calculate, no DB write)
    
    Body:
      {
        "xml_path": "C:\\...\\staging\\faktura_uploads\\invoice.xml"
      }
    
    Returns:
      {
        "header": {...},
        "items": [{...}],
        "matched_count": 4,
        "total_items": 5,
        "match_rate_pct": 80.0,
        "estimated_total": 15234.50
      }
    """
    try:
        data = request.get_json()
        xml_path = data.get('xml_path')
        
        if not xml_path or not os.path.exists(xml_path):
            return jsonify({"error": "Invalid xml_path"}), 400
        
        # Parse XML
        header, items = parse_invoice_xml(xml_path)
        
        # Load lookup for matching (use configured DB connection)
        try:
            base = os.path.dirname(os.path.dirname(__file__))
            cfg_path = os.path.join(base, 'configs', 'faktura_ai.json')
            cfg = load_config(cfg_path)
            sifra_set, barcode_set, _, _ = load_lookup_from_db(cfg['db'])
        except Exception:
            sifra_set, barcode_set = set(), set()
        
        # Match and calculate
        matched = 0
        total_calc = 0.0
        for item in items:
            sifra = item.get('sifra', '')
            barcode = item.get('barcode', '')
            if sifra in sifra_set or barcode in barcode_set:
                matched += 1
            
            qty = to_number(item.get('qty', '0'))
            price = to_number(item.get('price', '0'))
            rabat = to_number(item.get('rabat_pct', '0'))
            net_price = price * (1 - rabat/100.0)
            total_calc += qty * net_price
        
        match_rate = (matched / len(items) * 100) if items else 0
        
        return jsonify({
            "header": header,
            "items": items,
            "matched_count": matched,
            "total_items": len(items),
            "match_rate_pct": round(match_rate, 2),
            "estimated_total": round(total_calc, 2)
        })
    
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/faktura/execute", methods=["POST"])
def api_faktura_execute():
    """
    ⚡ Execute invoice import to ERP database (kalkopste + kalkstavke + kalkkasa)
    
    Body:
      {
        "xml_path": "C:\\...\\staging\\faktura_uploads\\invoice.xml",
        "magacin": "101",
        "periodid": 4,
        "userid": 14,
        "dry_run": false
      }
    
    Returns:
      {
        "success": true,
        "kalkid": 123,
        "broj": "45/24",
        "items_inserted": 5,
        "message": "Import completed successfully"
      }
    """
    try:
        if not SOPHARMA_ERP_AVAILABLE:
            return jsonify({"error": "sopharma_to_erp module not available"}), 500
        
        data = request.get_json()
        xml_path = data.get('xml_path')
        xml_filename = data.get('filename')  # optional, UI may send only the basename
        magacin = data.get('magacin', '101')
        periodid = int(data.get('periodid', 4))
        userid = int(data.get('userid', 14))
        dry_run = data.get('dry_run', False)
        
        # Resolve xml path(s) robustly. Accept single xml_path or a list xml_paths for bulk.
        resolved_paths = []
        if xml_path and os.path.exists(xml_path):
            resolved_path = xml_path
        else:
            # Try reconstructing from filename under staging dir
            base = os.path.dirname(os.path.dirname(__file__))
            staging_dir = os.path.join(base, 'staging', 'faktura_uploads')
            ensure_dir(staging_dir)
            candidates = []
            if xml_path:
                candidates.append(xml_path)
                candidates.append(os.path.join(staging_dir, os.path.basename(xml_path)))
                # Also try URL-decoded variant
                candidates.append(os.path.join(staging_dir, urllib.parse.unquote(os.path.basename(xml_path))))
            if xml_filename:
                candidates.append(os.path.join(staging_dir, xml_filename))
                candidates.append(os.path.join(staging_dir, urllib.parse.unquote(xml_filename)))
            for c in candidates:
                if c and os.path.exists(c):
                    resolved_path = c
                    break

        # After trying reconstruction, if xml_path present add to list
        if resolved_path and os.path.exists(resolved_path):
            resolved_paths.append(resolved_path)

        # Also accept bulk list in request
        if isinstance(data.get('xml_paths'), list):
            for p in data.get('xml_paths'):
                if p and os.path.exists(p):
                    resolved_paths.append(p)

        if not resolved_paths:
            return jsonify({"error": "Invalid xml_path(s)"}), 400
        
        # Parse XML (use sopharma parser for richer data)
        
        # Database config
        db_cfg = {
            'dbname': os.getenv('WPH_DB_NAME', 'wph_ai'),
            'user': os.getenv('WPH_DB_USER', 'postgres'),
            'password': os.getenv('WPH_PG18_PASS', ''),
            'host': os.getenv('WPH_DB_HOST', '127.0.0.1'),
            'port': int(os.getenv('WPH_DB_PORT', '5432')),
            'application_name': 'wphAI_web_faktura_import',
        }
        
        if not db_cfg['password']:
            return jsonify({"error": "WPH_PG18_PASS environment variable not set"}), 500
        
        # Connect and execute
        conn = psycopg2.connect(**db_cfg)
        try:
            allow_remote = os.getenv('WPH_WRITE_REMOTE', '0') == '1'
            
            kalk_id = insert_kalkulacija(
                conn, header, items, MP_CONFIG,
                dokvrsta='20',
                magacin=magacin,
                komintent='1',  # Will be auto-resolved from supplier name
                periodid=periodid,
                userid=userid,
                dry_run=dry_run,
                allow_remote_write=allow_remote
            )
            
            if kalk_id:
                # Move XML to archive after successful import
                if not dry_run:
                    base = os.path.dirname(os.path.dirname(__file__))
                    archive_dir = os.path.join(base, 'archive', 'faktura_imported', datetime.datetime.now().strftime('%Y%m%d'))
                    ensure_dir(archive_dir)
                    archive_path = os.path.join(archive_dir, os.path.basename(resolved_path))
                    os.rename(resolved_path, archive_path)
                
                return jsonify({
                    "success": True,
                    "kalkid": kalk_id,
                    "broj": header.get('broj_faktura', 'N/A'),
                    "items_inserted": len(items),
                    "message": "Import completed successfully" if not dry_run else "DRY-RUN: No data written"
                })
            else:
                return jsonify({"error": "insert_kalkulacija returned None"}), 500
        
        finally:
            conn.close()
    
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


# ============ Serbian eFaktura API Integration ============
SERBIA_BASE_URL = os.getenv('WPH_SERBIA_BASE_URL', 'https://api.efaktura.gov.rs')
SERBIA_API_KEY = os.getenv('WPH_SERBIA_API_KEY')

def serbia_api_request(endpoint, method='GET', params=None, data=None, json_data=None, raw=False):
    """Helper to make requests to Serbian eFaktura API"""
    if not SERBIA_API_KEY:
        # Return mock data for testing
        if '/purchase-invoice' in endpoint:
            return [
                {
                    "invoiceId": "12345",
                    "status": "Received",
                    "cirStatus": "Approved",
                    "comment": "Test invoice",
                    "lastModifiedUtc": "2025-11-15T10:00:00Z"
                },
                {
                    "invoiceId": "12346",
                    "status": "Pending",
                    "cirStatus": "InProgress",
                    "comment": "Another test",
                    "lastModifiedUtc": "2025-11-14T09:00:00Z"
                }
            ]
        return []
    
    url = f"{SERBIA_BASE_URL}{endpoint}"
    headers = {'Content-Type': 'application/json'}
    params = params or {}
    params['apiKey'] = SERBIA_API_KEY  # Assuming apiKey is query param
    print(f"[SERBIA API] Request {method} {url} params={params} headers={list(headers.keys())}")

    # Try primary method (apiKey as query param)
    params = params or {}
    params['apiKey'] = SERBIA_API_KEY
    try:
        response = requests.request(method, url, headers=headers, params=params, data=data, json=json_data, timeout=20)
        response.raise_for_status()
        if raw:
            return response.text
        return response.json()
    except requests.exceptions.HTTPError as e2:
        # first header fallback
        try:
            headers['x-api-key'] = SERBIA_API_KEY
            response = requests.request(method, url, headers=headers, params=params, data=data, json=json_data, timeout=20)
            response.raise_for_status()
            if raw:
                return response.text
            return response.json()
        except requests.exceptions.HTTPError:
            # next: ApiKey header
            try:
                headers.pop('x-api-key', None)
                headers['ApiKey'] = SERBIA_API_KEY
                response = requests.request(method, url, headers=headers, params=params, data=data, json=json_data, timeout=20)
                response.raise_for_status()
                if raw:
                    return response.text
                return response.json()
            except requests.exceptions.HTTPError:
                # Last resort: Bearer token
                headers.pop('ApiKey', None)
                headers['Authorization'] = f"Bearer {SERBIA_API_KEY}"
                response = requests.request(method, url, headers=headers, params=params, data=data, json=json_data, timeout=20)
                response.raise_for_status()
                if raw:
                    return response.text
                return response.json()
            try:
                # Try ApiKey header as used in Swagger and Public API examples
                headers.pop('x-api-key', None)
                headers['ApiKey'] = SERBIA_API_KEY
                response = requests.request(method, url, headers=headers, params=params, data=data, json=json_data)
                response.raise_for_status()
                if raw:
                    return response.text
                return response.json()
            except requests.exceptions.HTTPError as e3:
                try:
                    print(f"[SERBIA API] fallback ApiKey header failed: {e3.response.status_code} -> {e3.response.text[:400]}")
                except Exception:
                    print(f"[SERBIA API] fallback ApiKey header error: {str(e3)}")
            try:
                # Last resort: Bearer token
                headers.pop('x-api-key', None)
                headers['Authorization'] = f"Bearer {SERBIA_API_KEY}"
                response = requests.request(method, url, headers=headers, params=params, data=data, json=json_data)
                response.raise_for_status()
                return response.json()
            except Exception as e3:
                try:
                    print(f"[SERBIA API] fallback Bearer failed: {str(e3)}")
                except Exception:
                    pass
                # Re-raise original error for visibility
                raise e
    except requests.exceptions.ConnectionError as ce:
        # Common problem: default host in .env may not resolve on the current network.
        msg = str(ce)
        print(f"[SERBIA API] ConnectionError: {msg}")
        # Suggest a known working hostname if configuration appears unchanged
        if 'api.efaktura.gov.rs' in SERBIA_BASE_URL:
            print("[SERBIA API] Suggest overriding WPH_SERBIA_BASE_URL to https://efaktura.mfin.gov.rs if api.efaktura.gov.rs is not reachable.")
            # Try a common alternate host used in Swagger
            try_url = SERBIA_BASE_URL.replace('api.efaktura.gov.rs', 'efaktura.mfin.gov.rs') + endpoint
            print(f"[SERBIA API] Attempting fallback host: {try_url}")
            try:
                response = requests.request(method, try_url, headers=headers, params=params, data=data, json=json_data, timeout=20)
                response.raise_for_status()
                if raw:
                    return response.text
                return response.json()
            except Exception:
                pass
        raise
    except Exception:
        # Bubble up other unexpected exceptions
        raise

@app.route("/api/serbia/sales-invoices", methods=["GET"])
def api_serbia_sales_invoices():
    """
    📄 Get sales invoices from Serbian eFaktura
    
    Query params: page, size, dateFrom, dateTo, status, etc. (see swagger)
    """
    try:
        params = request.args.to_dict()
        data = serbia_api_request('/api/publicApi/sales-invoice', params=params)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/serbia/purchase-invoices", methods=["GET"])
def api_serbia_purchase_invoices():
    """
    📄 Get purchase invoices from Serbian eFaktura
    
    Query params: page, size, dateFrom, dateTo, status, etc.
    """
    try:
        # For now, return mock data since API key is not set
        if not SERBIA_API_KEY:
            mock_invoices = [
                {
                    "invoiceId": "12345",
                    "status": "Received",
                    "cirStatus": "Approved",
                    "comment": "Test invoice 1",
                    "lastModifiedUtc": "2025-11-15T10:00:00Z"
                },
                {
                    "invoiceId": "12346", 
                    "status": "Pending",
                    "cirStatus": "In Review",
                    "comment": "Test invoice 2",
                    "lastModifiedUtc": "2025-11-14T15:30:00Z"
                }
            ]
            return jsonify(mock_invoices)
        
        params = request.args.to_dict()
        # Use overview endpoint for listing summaries
        try:
            data = serbia_api_request('/api/publicApi/purchase-invoice/overview', params=params)
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            print(f"[SERBIA API ERROR] purchase-invoices overview failed: {e}\n{tb}")
            # Try fallback endpoint to get additional info
            try:
                data = serbia_api_request('/api/publicApi/purchase-invoice', params=params)
            except Exception as e2:
                # If both fail, return the error text with status 502
                err_msg = str(e2)
                if hasattr(e2, 'response') and e2.response is not None:
                    try:
                        err_body = e2.response.text
                    except Exception:
                        err_body = None
                else:
                    err_body = None
                return jsonify({"error": "efaktura overview failed", "message": err_msg, "body": err_body}), 502
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/serbia/companies", methods=["GET"])
def api_serbia_companies():
    """
    🏢 Get all registered companies from Serbian eFaktura
    """
    try:
        data = serbia_api_request('/api/publicApi/getAllCompanies')
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/serbia/vat-recordings/group", methods=["GET"])
def api_serbia_vat_group():
    """
    💰 Get group VAT recordings from Serbian eFaktura
    """
    try:
        params = request.args.to_dict()
        data = serbia_api_request('/api/publicApi/vat-recording/group', params=params)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/serbia/fetch-efaktura", methods=["POST"])
def api_serbia_fetch_efaktura():
    """
    ⬇️ Fetch invoices from Serbian eFaktura API into staging.

    Body (optional): {"days": 7} or {"from": "YYYY-MM-DD", "to": "YYYY-MM-DD"}
    """
    try:
        data = request.get_json(silent=True) or {}
        if 'from' in data and 'to' in data:
            dfrom = datetime.datetime.fromisoformat(data['from']).date()
            dto = datetime.datetime.fromisoformat(data['to']).date()
        else:
            days = int(data.get('days', 7))
            dto = datetime.date.today()
            dfrom = dto - datetime.timedelta(days=days)

        # Keep XMLs under the canonical staging dir used by Faktura AI
        base = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'staging', 'faktura_uploads')
        ensure_dir(base)
        # If API key not set, return 200 with empty list (do not break UI)
        if not SERBIA_API_KEY:
            return jsonify({"fetched": 0, "paths": [], "note": "WPH_SERBIA_API_KEY not set - no fetch performed"})

        params = request.get_json(silent=True) or {}
        page_size = int(params.get('size', 50))
        max_pages = int(params.get('max_pages', 20))

        # Build the search params (dates, statuses) to pass through
        search_params = {}
        # support both days and explicit from/to
        if 'from' in params and 'to' in params:
            search_params['dateFrom'] = params['from']
            search_params['dateTo'] = params['to']
        else:
            days = int(params.get('days', 7))
            dto = datetime.date.today()
            dfrom = dto - datetime.timedelta(days=days)
            search_params['dateFrom'] = dfrom.isoformat()
            search_params['dateTo'] = dto.isoformat()

        # page through results
        page = 0
        fetched_paths = []
        errors = []
        while page < max_pages:
            page_params = search_params.copy()
            page_params.update({'page': page, 'size': page_size})
            try:
                data = serbia_api_request('/api/publicApi/purchase-invoice/overview', params=page_params)
            except Exception as e:
                # Try fallback host suggestion to help debugging (common DNS mismatch)
                errors.append(str(e))
                break

            # determine invoice items list
            if isinstance(data, dict):
                items = data.get('content') or data.get('items') or data.get('invoices') or data.get('purchaseInvoices') or []
            elif isinstance(data, list):
                items = data
            else:
                items = []

            if not items:
                break

            for item in items:
                # try common keys for invoiceId
                inv_id = item.get('invoiceId') or item.get('purchaseInvoiceId') or item.get('id') or item.get('invoiceNumber')
                if not inv_id:
                    continue
                try:
                    xml_text = serbia_api_request('/api/publicApi/purchase-invoice/xml', params={'invoiceId': inv_id}, raw=True)
                except Exception as e:
                    # record and continue with other invoices
                    errors.append(f"invoiceId={inv_id}: {str(e)}")
                    continue

                # Save XML to staging - keep unique names
                ts = datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S')
                safe_id = ''.join([c for c in str(inv_id) if c.isalnum() or c in ('-', '_')])
                filename = f"efaktura_{safe_id}_{ts}.xml"
                target = os.path.join(base, filename)
                # If file exists, add a counter
                i = 1
                while os.path.exists(target):
                    target = os.path.join(base, f"efaktura_{safe_id}_{ts}_{i}.xml")
                    i += 1
                with open(target, 'w', encoding='utf-8') as fh:
                    fh.write(xml_text)
                fetched_paths.append(target)

            # break out if we received fewer emails than page_size
            if len(items) < page_size:
                break
            page += 1

        return jsonify({"fetched": len(fetched_paths), "paths": fetched_paths, "errors": errors})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500

@app.route("/api/serbia/purchase-invoice/xml/<invoiceId>", methods=["GET"])
def api_serbia_purchase_invoice_xml(invoiceId):
    """
    📄 Get purchase invoice XML from Serbian API
    """
    try:
        # Fetch real XML from eFaktura public API when API key is set
        if SERBIA_API_KEY:
            try:
                # The eFaktura public API uses query param invoiceId
                xml_text = serbia_api_request('/api/publicApi/purchase-invoice/xml', params={'invoiceId': invoiceId}, raw=True)
                return Response(xml_text, mimetype='application/xml')
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500

        # Mock XML (fallback)
        xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2">
  <cbc:ID>{invoiceId}</cbc:ID>
  <cbc:IssueDate>2025-11-16</cbc:IssueDate>
  <!-- Mock XML content -->
</Invoice>"""
        return Response(xml_content, mimetype='application/xml')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/serbia/preview/<invoiceId>", methods=["GET"])
def api_serbia_preview(invoiceId):
    """
    👁️ Preview Serbian invoice before import
    """
    try:
        # Mock preview data - in reality, fetch XML and parse
        data = {
            "header": {
                "invoice_no": invoiceId,
                "supplier": "Mock Serbian Supplier",
                "invoice_date": "2025-11-16",
                "currency": "RSD"
            },
            "items": [
                {
                    "sifra": "12345",
                    "naziv": "Mock Item 1",
                    "qty": 10,
                    "unit_cost": 100.0,
                    "line_total": 1000.0
                }
            ],
            "validation": {
                "total_calc": 1000.0,
                "status": "MOCK"
            }
        }
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/serbia/execute/<invoiceId>", methods=["POST"])
def api_serbia_execute(invoiceId):
    """
    ⚡ Execute import of Serbian invoice to ERP
    """
    try:
        # Mock execute - in reality, save XML and run faktura_import.py
        data = {
            "status": "success",
            "invoice_id": invoiceId,
            "message": "Mock import completed"
        }
        return jsonify(data)
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500

# Add more endpoints as needed...


if __name__ == "__main__":
    app.logger.info("Starting backend WPH AI app: UI_PUBLIC_PATH=%s, USE_DB=%s", UI_PUBLIC_PATH or 'None', USE_DB)
    app.run(host="0.0.0.0", port=int(os.getenv("APP_PORT","8055")), debug=False, use_reloader=False)